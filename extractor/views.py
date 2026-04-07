import os
import csv
import traceback
from django.forms import ValidationError
import pandas as pd
import zipfile
from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.db.models import Q, Count
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import ExcelData, Cliente ,TipoServicio, Proyecto, Ticket, Usuario
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseServerError, JsonResponse
from django.db import models
from openpyxl import Workbook
from django.core.paginator import Paginator
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from django.utils import timezone
from .models import SolicitudPruebas 
from io import BytesIO
import logging
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from .models import Cliente, Proyecto, TipoServicio, Ticket, ExcelData
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.shortcuts import redirect
from django.conf import settings
from django.core.cache import cache
from .forms import RegistroUsuarioForm
import json
from django.views.decorators.csrf import csrf_exempt

logger = logging.getLogger(__name__)

# ===== VISTAS PÚBLICAS (NO requieren login) =====
def login_view(request):
    """Vista personalizada de login"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            next_url = request.GET.get('next', 'extractor:upload_excel')
            return redirect(next_url)
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    return render(request, 'extractor/login.html')

def logout_view(request):
    """Vista personalizada de logout"""
    logout(request)
    return redirect('extractor:login')

@login_required
def export_clientes_csv(request):
    """
    Exporta clientes a CSV
    """
    try:
        # Obtener datos
        clientes = Cliente.objects.all()
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))  # BOM para Excel
        
        # Nombre del archivo
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"clientes_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Crear writer
        writer = csv.writer(response)
        
        # Escribir encabezados
        writer.writerow(['ID', 'Nombre', 'Nomenclatura', 'Activo', 'Fecha Creación'])
        
        # Escribir datos
        for cliente in clientes:
            writer.writerow([
                cliente.id,
                cliente.nombre,
                cliente.nomenclatura,
                'Sí' if cliente.activo else 'No',
                cliente.fecha_creacion.strftime('%d/%m/%Y %H:%M') if cliente.fecha_creacion else ''
            ])
        
        logger.info(f"Usuario {request.user} exportó clientes - {clientes.count()} registros")
        return response
        
    except Exception as e:
        logger.error(f"Error exportando clientes: {str(e)}", exc_info=True)
        messages.error(request, "Error al exportar clientes")
        return redirect('extractor:clientes_list')

@login_required
def export_proyectos_csv(request):
    """
    Exporta proyectos a CSV
    """
    try:
        proyectos = Proyecto.objects.all().select_related('cliente')
        
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))
        
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"proyectos_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Cliente', 'Nombre', 'Código', 'Descripción', 'Activo', 'Fecha Inicio', 'Fecha Fin'])
        
        for proyecto in proyectos:
            writer.writerow([
                proyecto.id,
                proyecto.cliente.nombre if proyecto.cliente else '',
                proyecto.nombre,
                proyecto.codigo,
                proyecto.descripcion or '',
                'Sí' if proyecto.activo else 'No',
                proyecto.fecha_inicio.strftime('%d/%m/%Y') if proyecto.fecha_inicio else '',
                proyecto.fecha_fin.strftime('%d/%m/%Y') if proyecto.fecha_fin else ''
            ])
        
        logger.info(f"Usuario {request.user} exportó proyectos - {proyectos.count()} registros")
        return response
        
    except Exception as e:
        logger.error(f"Error exportando proyectos: {str(e)}", exc_info=True)
        messages.error(request, "Error al exportar proyectos")
        return redirect('extractor:proyectos_list')
    
@login_required
def export_tipos_servicio_csv(request):
    """
    Exporta tipos de servicio a CSV
    """
    try:
        tipos = TipoServicio.objects.all()
        
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))
        
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"tipos_servicio_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Nombre', 'Nomenclatura', 'Activo', 'Fecha Creación'])
        
        for tipo in tipos:
            writer.writerow([
                tipo.id,
                tipo.nombre,
                tipo.nomenclatura,
                'Sí' if tipo.activo else 'No',
                tipo.fecha_creacion.strftime('%d/%m/%Y %H:%M') if tipo.fecha_creacion else ''
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exportando tipos de servicio: {str(e)}", exc_info=True)
        messages.error(request, "Error al exportar tipos de servicio")
        return redirect('tipos_servicio_list')

@login_required
def export_tickets_csv_view(request):
    """
    Exporta tickets a CSV (versión mejorada de export_tickets_excel pero en CSV)
    """
    try:
        # Usar los mismos filtros que en ticket_list
        tickets = Ticket.objects.all().select_related('cliente', 'proyecto', 'tipo_servicio', 'excel_data')
        
        # Aplicar filtros si vienen en GET
        estado = request.GET.get('estado')
        cliente_id = request.GET.get('cliente')
        proyecto_id = request.GET.get('proyecto')
        
        if estado:
            tickets = tickets.filter(estado=estado)
        if cliente_id:
            tickets = tickets.filter(cliente_id=cliente_id)
        if proyecto_id:
            tickets = tickets.filter(proyecto_id=proyecto_id)
        
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))
        
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"tickets_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Código Ticket', 'Estado', 'Cliente', 'Proyecto', 
            'Tipo Servicio', 'Responsable', 'Líder Proyecto', 'Versión',
            'Fecha Creación'
        ])
        
        for ticket in tickets:
            writer.writerow([
                ticket.id,
                ticket.codigo,
                ticket.get_estado_display(),
                ticket.cliente.nombre if ticket.cliente else '',
                ticket.proyecto.nombre if ticket.proyecto else '',
                ticket.tipo_servicio.nombre if ticket.tipo_servicio else '',
                ticket.responsable_solicitud or '',
                ticket.lider_proyecto or '',
                ticket.numero_version or '',
                ticket.fecha_creacion.strftime('%d/%m/%Y %H:%M')
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exportando tickets: {str(e)}", exc_info=True)
        messages.error(request, "Error al exportar tickets")
        return redirect('extractor:ticket_list')
    

def extract_excel_data(file_path):
    """
    Extrae las celdas específicas según las reglas dadas
    AHORA SOPORTA NOMBRES DIRECTOS (no solo IDs)
    """
    try:
        # Verificar que la hoja existe
        try:
            df = pd.read_excel(file_path, sheet_name='Solicitud de Pruebas V4', header=None)
        except ValueError as e:
            if "No sheet named" in str(e):
                raise Exception("El archivo no contiene la hoja 'Solicitud de Pruebas V4'")
            else:
                raise Exception(f"Error al leer el archivo: {str(e)}")
        
        # Inicializar diccionario para datos
        extracted_data = {}
        
        # Función auxiliar para limpiar valores
        def clean_value(value):
            """Limpia el valor extraído"""
            if pd.isna(value):
                return ""
            return str(value).strip()
        
        # Extraer CLIENTE (C5) - AHORA PUEDE SER NOMBRE O ID
        try:
            cliente_valor = clean_value(df.iat[4, 2])
            extracted_data['cliente'] = cliente_valor
            print(f"📌 Cliente extraído: '{cliente_valor}'")
        except:
            extracted_data['cliente'] = ""
        
        # Extraer PROYECTO (H5) - AHORA PUEDE SER NOMBRE O ID
        try:
            proyecto_valor = clean_value(df.iat[4, 7])
            extracted_data['proyecto'] = proyecto_valor
            print(f"📌 Proyecto extraído: '{proyecto_valor}'")
        except:
            extracted_data['proyecto'] = ""
        
        # Extraer TIPO DE PRUEBAS (D8) - AHORA PUEDE SER NOMBRE O ID
        try:
            tipo_pruebas_valor = clean_value(df.iat[7, 3])
            extracted_data['tipo_pruebas'] = tipo_pruebas_valor
            print(f"📌 Tipo de Pruebas extraído: '{tipo_pruebas_valor}'")
        except:
            extracted_data['tipo_pruebas'] = ""
        
        # Extraer responsable_solicitud (D12)
        try:
            extracted_data['responsable_solicitud'] = clean_value(df.iat[11, 3])
        except:
            extracted_data['responsable_solicitud'] = ""
        
        # Extraer lider_proyecto (J12)
        try:
            extracted_data['lider_proyecto'] = clean_value(df.iat[11, 9])
        except:
            extracted_data['lider_proyecto'] = ""
        
        # Extraer tipo_aplicacion (D17)
        try:
            extracted_data['tipo_aplicacion'] = clean_value(df.iat[16, 3])
        except:
            extracted_data['tipo_aplicacion'] = ""
        
        # Extraer numero_version (M17)
        try:
            extracted_data['numero_version'] = clean_value(df.iat[16, 12])
        except:
            extracted_data['numero_version'] = ""
        
        # Extraer funcionalidad_liberacion (D20)
        try:
            funcionalidad = clean_value(df.iat[19, 3])
            if pd.notna(df.iat[20, 3]):
                funcionalidad += "\n" + clean_value(df.iat[20, 3])
            extracted_data['funcionalidad_liberacion'] = funcionalidad
        except:
            extracted_data['funcionalidad_liberacion'] = ""
        
        # Extraer detalle_cambios (a partir de D22)
        try:
            detalle_cambios = ""
            row = 21  # Fila 22 (0-indexed)
            while row < 30 and pd.notna(df.iat[row, 3]):
                cell_value = clean_value(df.iat[row, 3])
                # Ignorar si es el texto de encabezado repetido
                if "📝 Descripción de Cambios" not in cell_value and "Funcionalidad de la liberación:" not in cell_value:
                    detalle_cambios += cell_value + "\n"
                row += 1
            extracted_data['detalle_cambios'] = detalle_cambios.strip()
        except:
            extracted_data['detalle_cambios'] = ""
        
        # 🔧 CORRECCIÓN: Extraer justificacion_cambio (fila 24)
        try:
            justificacion = ""
            
            # Buscar la fila de justificación - en tu Excel está en fila 23 (0-indexed)
            # Observando tu archivo, la justificación está en D24 (row=23)
            if pd.notna(df.iat[23, 3]):  # D24
                cell_value = clean_value(df.iat[23, 3])
                # Ignorar si es el texto de encabezado
                if "📝 Descripción de Cambios" not in cell_value and "Funcionalidad de la liberación:" not in cell_value:
                    justificacion = cell_value
            
            # Si no hay texto en D24, intentar buscar por el encabezado "Justificación"
            if not justificacion:
                justificacion_row = None
                for row in range(21, 30):
                    if pd.notna(df.iat[row, 2]) and "Justificación" in str(df.iat[row, 2]):
                        justificacion_row = row
                        break
                
                if justificacion_row is not None:
                    content_row = justificacion_row + 1
                    while content_row < 40 and pd.notna(df.iat[content_row, 3]):
                        justificacion += clean_value(df.iat[content_row, 3]) + "\n"
                        content_row += 1
            
            extracted_data['justificacion_cambio'] = justificacion.strip()
            
        except Exception as e:
            print(f"⚠️ Error extrayendo justificación: {e}")
            extracted_data['justificacion_cambio'] = ""
        
        # DEPURACIÓN: Mostrar valores extraídos
        print("\n=== VALORES EXTRAÍDOS DEL EXCEL ===")
        for key, value in extracted_data.items():
            print(f"{key}: '{value}'")
        print("=====================================\n")
        
        return extracted_data
        
    except Exception as e:
        print(f"❌ Error al extraer datos: {e}")
        raise

@login_required
def upload_excel(request):
    """
    Procesa la carga de archivos Excel y genera tickets
    CORREGIDO: Validación de nombre de archivo con SolicitudPruebas
    """
    if request.method == 'POST':
        # Inicializar variables
        fs = FileSystemStorage()
        file_path = None
        filename = None
        
        try:
            # Obtener datos del formulario
            tipo_servicio_form = request.POST.get('tipo_servicio', '').strip()
            excel_file = request.FILES.get('excel_file')
            
            # ===== VALIDACIONES INICIALES =====
            if not tipo_servicio_form:
                messages.error(request, 'Por favor selecciona un tipo de servicio')
                return render(request, 'extractor/upload.html')
            
            if not excel_file:
                messages.error(request, 'Por favor selecciona un archivo Excel')
                return render(request, 'extractor/upload.html')
            
            # Validar extensión del archivo
            allowed_extensions = ['.xlsx', '.xls']
            file_extension = os.path.splitext(excel_file.name)[1].lower()
            
            if file_extension not in allowed_extensions:
                messages.error(request, 'Formato de archivo no válido. Solo se permiten archivos .xlsx y .xls')
                return render(request, 'extractor/upload.html')
            
            # ===== VALIDAR NOMBRE DE ARCHIVO CON SOLICITUDES =====
            nombre_archivo_subido = excel_file.name
            print(f"📁 Nombre de archivo subido: {nombre_archivo_subido}")
            
            # Buscar solicitud que coincida con el nombre del archivo
            solicitud = SolicitudPruebas.objects.filter(
                nombre_archivo=nombre_archivo_subido,
                tiene_ticket=False  # Solo buscar solicitudes sin ticket
            ).first()
            
            if solicitud:
                print(f"✅ Solicitud encontrada: ID {solicitud.id}")
                solicitud_encontrada = True
                # Guardar referencia de la solicitud para usarla después
            else:
                print(f"❌ No se encontró solicitud con nombre: {nombre_archivo_subido}")
                solicitud_encontrada = False
                solicitud = None
            
            # ===== GUARDAR ARCHIVO TEMPORAL =====
            filename = fs.save(excel_file.name, excel_file)
            file_path = os.path.join(settings.MEDIA_ROOT, filename)
            
            # Verificar que el archivo se guardó correctamente
            if not os.path.exists(file_path):
                raise Exception("No se pudo guardar el archivo temporal")
            
            print(f"📁 Archivo temporal guardado: {file_path}")
            
            # ===== PROCESAR EL ARCHIVO =====
            # Extraer datos del Excel
            extracted_data = extract_excel_data(file_path)
            
            # Validar campos obligatorios
            campos_obligatorios = ['cliente', 'proyecto', 'tipo_pruebas']
            campos_faltantes = [campo for campo in campos_obligatorios if not extracted_data.get(campo)]
            
            if campos_faltantes:
                mensaje_error = "❌ El archivo no contiene los siguientes campos obligatorios:\n"
                mensaje_error += "\n".join(f"• {campo}" for campo in campos_faltantes)
                raise ValidationError(mensaje_error)
            
            # Buscar objetos por nombre o ID
            cliente_obj = find_object_by_name_or_id(Cliente, extracted_data.get('cliente', ''), 'nombre')
            proyecto_obj = find_object_by_name_or_id(Proyecto, extracted_data.get('proyecto', ''), 'nombre')
            tipo_prueba_obj = find_object_by_name_or_id(TipoServicio, extracted_data.get('tipo_pruebas', ''), 'nombre')
            
            # Validar objetos encontrados
            objetos_no_encontrados = []
            if not cliente_obj:
                objetos_no_encontrados.append(f"Cliente '{extracted_data.get('cliente', '')}'")
            if not proyecto_obj:
                objetos_no_encontrados.append(f"Proyecto '{extracted_data.get('proyecto', '')}'")
            if not tipo_prueba_obj:
                objetos_no_encontrados.append(f"Tipo de Pruebas '{extracted_data.get('tipo_pruebas', '')}'")
            
            if objetos_no_encontrados:
                mensaje_error = "❌ No se encontraron en el catálogo:\n"
                mensaje_error += "\n".join(f"• {objeto}" for objeto in objetos_no_encontrados)
                raise ValidationError(mensaje_error)
            
            # Validar que el proyecto pertenezca al cliente
            if proyecto_obj.cliente_id != cliente_obj.id:
                raise ValidationError(f'❌ El proyecto "{proyecto_obj.nombre}" no pertenece al cliente "{cliente_obj.nombre}"')
            
            # ===== GENERAR TICKET =====
            # Preparar nomenclaturas
            nomenclaturas = {
                'cliente_nomenclatura': cliente_obj.nomenclatura,
                'proyecto_nomenclatura': proyecto_obj.codigo,
                'tipo_pruebas_nomenclatura': tipo_prueba_obj.nomenclatura,
                'tipo_servicio_nomenclatura': tipo_servicio_form
            }
            
            objetos_encontrados = {
                'cliente_obj': cliente_obj,
                'proyecto_obj': proyecto_obj,
                'tipo_servicio_obj': tipo_prueba_obj
            }
            
            # Generar ticket
            ticket_code, ticket_obj = generate_and_save_ticket(
                extracted_data=extracted_data,
                tipo_servicio_form=tipo_servicio_form,
                nomenclaturas=nomenclaturas,
                objetos_encontrados=objetos_encontrados,
                request=request
            )
            
            # ===== NUEVO: CREAR INCIDENCIA EN JIRA =====
            if ticket_obj:
                try:
                    from .jira_helper import JiraClient
                    from django.utils import timezone
                    
                    # Preparar datos para Jira
                    jira_data = {
                        'codigo': ticket_code,
                        'cliente': cliente_obj.nombre,
                        'proyecto': proyecto_obj.nombre,
                        'tipo_servicio': tipo_servicio_form,
                        'responsable_solicitud': extracted_data.get('responsable_solicitud', ''),
                        'lider_proyecto': extracted_data.get('lider_proyecto', ''),
                        'numero_version': extracted_data.get('numero_version', ''),
                        'funcionalidad_liberacion': extracted_data.get('funcionalidad_liberacion', ''),
                        'detalle_cambios': extracted_data.get('detalle_cambios', ''),
                        'justificacion_cambio': extracted_data.get('justificacion_cambio', ''),
                        'fecha': timezone.now().strftime('%d/%m/%Y %H:%M'),
                        'usuario': request.user.username if request.user.is_authenticated else 'Sistema',
                    }
                    
                    # Inicializar cliente Jira
                    jira_client = JiraClient()
                    
                    # Crear incidencia
                    jira_issue = jira_client.create_issue(jira_data)
                    
                    if jira_issue:
                        # Guardar la información de Jira en el ticket
                        ticket_obj.jira_issue_key = jira_issue.key
                        ticket_obj.jira_issue_url = jira_issue.permalink()
                        ticket_obj.fecha_sincronizacion_jira = timezone.now()
                        ticket_obj.save()
                        
                        # Mensaje para el usuario
                        messages.info(
                            request, 
                            f'📋 Incidencia creada en Jira: {jira_issue.key}'
                        )
                        
                        print(f"✅ Ticket vinculado a Jira issue: {jira_issue.key}")
                    else:
                        print("⚠️ No se pudo crear incidencia en Jira")
                        
                except Exception as jira_error:
                    print(f"⚠️ Error en integración Jira: {jira_error}")
                    # No interrumpimos el flujo principal
                    pass
            
            # Guardar en la base de datos ExcelData
            excel_data = ExcelData.objects.create(
                cliente=str(cliente_obj.id),
                proyecto=str(proyecto_obj.id),
                tipo_pruebas=str(tipo_prueba_obj.id),
                tipo_servicio=tipo_servicio_form,
                responsable_solicitud=extracted_data.get('responsable_solicitud', ''),
                lider_proyecto=extracted_data.get('lider_proyecto', ''),
                tipo_aplicacion=extracted_data.get('tipo_aplicacion', ''),
                numero_version=extracted_data.get('numero_version', ''),
                funcionalidad_liberacion=extracted_data.get('funcionalidad_liberacion', ''),
                detalle_cambios=extracted_data.get('detalle_cambios', ''),
                justificacion_cambio=extracted_data.get('justificacion_cambio', ''),
                ticket_code=ticket_code
            )
            
            # Asociar el ticket con los datos del Excel
            if ticket_obj:
                ticket_obj.excel_data = excel_data
                ticket_obj.save()
            
            # ===== SI SE ENCONTRÓ UNA SOLICITUD, ACTUALIZARLA =====
            if solicitud_encontrada and solicitud:
                # Asociar el ticket a la solicitud
                solicitud.ticket = ticket_obj
                solicitud.tiene_ticket = True
                solicitud.fecha_asociacion_ticket = timezone.now()
                solicitud.save()
                
                print(f"✅ Solicitud ID {solicitud.id} actualizada: ticket asociado {ticket_code}")
            
            # ===== ÉXITO =====
            if solicitud_encontrada:
                messages.success(request, f'✅ Archivo procesado exitosamente. Ticket generado y asociado a solicitud #{solicitud.id}. Ticket: {ticket_code}')
            else:
                messages.success(request, f'✅ Archivo procesado exitosamente. Ticket generado: {ticket_code}')
            
            # Preparar datos para la plantilla
            data_for_template = {
                'cliente': excel_data.cliente,
                'proyecto': excel_data.proyecto,
                'tipo_pruebas': excel_data.tipo_pruebas,
                'tipo_servicio': excel_data.tipo_servicio,
                'responsable_solicitud': excel_data.responsable_solicitud,
                'lider_proyecto': excel_data.lider_proyecto,
                'tipo_aplicacion': excel_data.tipo_aplicacion,
                'numero_version': excel_data.numero_version,
                'funcionalidad_liberacion': excel_data.funcionalidad_liberacion,
                'detalle_cambios': excel_data.detalle_cambios,
                'justificacion_cambio': excel_data.justificacion_cambio,
                'extracted_date': excel_data.extracted_date
            }
            
            ticket_parts = generate_ticket_parts(ticket_code)
            
            return redirect('extractor:ticket_detail', id=ticket_obj.id)
            
        except ValidationError as e:
            # Errores esperados (validación de usuario)
            messages.error(request, str(e))
            return render(request, 'extractor/upload.html')
            
        except Exception as e:
            # Errores inesperados
            print(f"❌ ERROR en procesamiento: {str(e)}")
            import traceback
            traceback.print_exc()
            
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return render(request, 'extractor/upload.html')
            
        finally:
            # ===== LIMPIEZA GARANTIZADA =====
            # Este bloque SIEMPRE se ejecuta, haya o no error
            if file_path and os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    print(f"✅ Archivo temporal eliminado: {file_path}")
                except Exception as e:
                    print(f"⚠️ Error al eliminar archivo temporal: {e}")
    
    # GET request
    return render(request, 'extractor/upload.html')

def find_object_by_name_or_id(model, value, field_name="nombre"):
    """
    Busca un objeto por nombre o ID
    """
    if not value or value == "":
        return None
    
    value_str = str(value).strip()
    
    # Intentar 1: Buscar por ID (si es un número)
    try:
        id_value = int(float(value_str))  # Convertir a float primero por si viene "1.0"
        obj = model.objects.filter(id=id_value).first()
        if obj:
            print(f"✅ Encontrado por ID: {model.__name__} ID={id_value} → {obj}")
            return obj
    except (ValueError, TypeError):
        pass  # No es un número, continuar con búsqueda por nombre
    
    # Intentar 2: Buscar por nombre exacto
    obj = model.objects.filter(**{field_name: value_str}).first()
    if obj:
        print(f"✅ Encontrado por nombre exacto: {model.__name__} '{value_str}' → {obj}")
        return obj
    
    # Intentar 3: Buscar por nombre que contenga (case insensitive)
    filter_kwargs = {f"{field_name}__icontains": value_str}
    obj = model.objects.filter(**filter_kwargs).first()
    if obj:
        print(f"✅ Encontrado por nombre que contiene: {model.__name__} '{value_str}' → {obj}")
        return obj
    
    # Intentar 4: Buscar por nomenclatura
    if hasattr(model, 'nomenclatura'):
        obj = model.objects.filter(nomenclatura=value_str).first()
        if obj:
            print(f"✅ Encontrado por nomenclatura: {model.__name__} '{value_str}' → {obj}")
            return obj
    
    print(f"❌ No encontrado: {model.__name__} con valor '{value_str}'")
    return None

# Añade esta función para generar el código del ticket
def generate_ticket_code(extracted_data, tipo_servicio):
    """Genera el código del ticket basado en los datos"""
    # Aquí puedes implementar tu lógica para generar el código del ticket
    # Ejemplo: BID-PRU-F&REG-10-TEL-OTR-001
    
    cliente_nom = "TEL"  # Obtener de la base de datos
    proyecto_nom = "OTR"  # Obtener de la base de datos
    version = extracted_data.get('numero_version', '10')
    
    # Determinar el código del tipo de servicio
    tipo_servicio_code = tipo_servicio  # Ya viene del formulario: PRU, EST, G&A
    
    consecutivo = "001"  # Deberías obtener este de la base de datos (último + 1)
    
    return f"BID-{tipo_servicio_code}-F&REG-{version}-{cliente_nom}-{proyecto_nom}-{consecutivo}"

@login_required
def data_list(request):
    data = ExcelData.objects.all().order_by('-extracted_date')
    return render(request, 'extractor/list.html', {'data_list': data})

@login_required
def clientes_list(request):
    try:
        clientes = Cliente.objects.all()
        
        # Debug: imprime los parámetros GET
        print(f"GET parameters: {request.GET}")
        
        # Ordenamiento
        orden = request.GET.get('orden', 'id')
        print(f"Orden solicitado: {orden}")
        
        # Diccionario de ordenamiento permitido
        orden_permitido = {
            'id': 'id', 
            '-id': '-id',
            'nomenclatura': 'nomenclatura', 
            '-nomenclatura': '-nomenclatura',
            'nombre': 'nombre', 
            '-nombre': '-nombre',
            'activo': 'activo', 
            '-activo': '-activo',
            'fecha_creacion': 'fecha_creacion', 
            '-fecha_creacion': '-fecha_creacion',
        }
        
        orden_final = orden_permitido.get(orden, 'id')
        print(f"Orden final: {orden_final}")
        
        clientes = clientes.order_by(orden_final)
        print(f"Query SQL: {clientes.query}")
        
        # Anotar con conteo de proyectos
        clientes = clientes.annotate(
            total_proyectos=Count('proyectos')
        )
        
        context = {
            'clientes': clientes,
        }
        return render(request, 'catalogos/clientes_list.html', context)
        
    except Exception as e:
        # Capturar el error completo
        error_traceback = traceback.format_exc()
        print(f"ERROR EN clientes_list: {str(e)}")
        print(f"Traceback: {error_traceback}")
        
        # Devolver el error en la respuesta para verlo en el navegador
        return HttpResponse(f"""
            <h1>Error en clientes_list</h1>
            <p><strong>Error:</strong> {str(e)}</p>
            <h2>Traceback:</h2>
            <pre>{error_traceback}</pre>
            <h2>GET parameters:</h2>
            <pre>{dict(request.GET)}</pre>
        """)

@login_required
def cliente_create(request):
    """Crear un nuevo cliente"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        nomenclatura = request.POST.get('nomenclatura', '').strip().upper()
        
        # Validaciones
        if not nombre or not nomenclatura:
            messages.error(request, 'Todos los campos son obligatorios')
            return render(request, 'catalogos/cliente_form.html')
        
        if len(nomenclatura) > 5:
            messages.error(request, 'La nomenclatura no puede tener más de 5 caracteres')
            return render(request, 'catalogos/cliente_form.html')
        
        # Verificar si ya existe la nomenclatura
        if Cliente.objects.filter(nomenclatura=nomenclatura).exists():
            messages.error(request, f'La nomenclatura "{nomenclatura}" ya existe')
            return render(request, 'catalogos/cliente_form.html')
        
        try:
            cliente = Cliente.objects.create(
                nombre=nombre,
                nomenclatura=nomenclatura,
                activo=request.POST.get('activo', 'on') == 'on'
            )
            messages.success(request, f'Cliente "{cliente.nombre}" creado exitosamente')
            return redirect('extractor:clientes_list')
        except Exception as e:
            messages.error(request, f'Error al crear cliente: {str(e)}')
    
    return render(request, 'catalogos/cliente_form.html')


def get_next_consecutivo(tipo_servicio_code, tipo_pruebas_nom, tipo_pruebas_id, cliente_nom, proyecto_nom):
    """Obtiene el siguiente número consecutivo para tickets con los mismos datos"""
    try:
        # Parámetros de búsqueda - CORREGIDO
        filtro = {
            'empresa_code': "BID",
            'tipo_servicio_code': tipo_servicio_code,
            'funcion_code': tipo_pruebas_nom,  # ← Esto es la nomenclatura
            'version_code': tipo_pruebas_id,    # ← Esto es el ID (valor numérico)
            'cliente_code': cliente_nom,
            'proyecto_code': proyecto_nom
        }
        
        print(f"🔍 Buscando tickets similares con filtro: {filtro}")
        
        # Buscar TODOS los tickets con los MISMOS parámetros
        tickets_similares = Ticket.objects.filter(**filtro)
        
        print(f"📊 Tickets encontrados: {tickets_similares.count()}")
        
        if tickets_similares.exists():
            for ticket in tickets_similares:
                print(f"   - {ticket.codigo} (consecutivo: {ticket.consecutivo})")
            
            max_consecutivo = tickets_similares.aggregate(models.Max('consecutivo'))['consecutivo__max']
            print(f"🎯 Máximo consecutivo encontrado: {max_consecutivo}")
            
            siguiente = max_consecutivo + 1
            print(f"🔄 Siguiente consecutivo: {siguiente}")
            return siguiente
        else:
            print(f"✨ No hay tickets similares, empezando en 1")
            return 1
    except Exception as e:
        print(f"⚠️ Error al obtener consecutivo: {str(e)}")
        traceback.print_exc()  # ← Añadir traceback completo
        return 1


def generate_ticket_parts(ticket_code):
    """Divide el código del ticket en partes para mostrar en el desglose"""
    parts = ticket_code.split('-')
    
    # Asegurar que tengamos 7 partes
    if len(parts) < 7:
        # Rellenar con valores por defecto si faltan partes
        default_parts = ['BID', 'PRU', 'F&REG', '10', 'TEL', 'OTR', '001']
        for i in range(7):
            if i >= len(parts) or not parts[i]:
                parts.append(default_parts[i])
    
    return parts

@login_required
def cliente_edit(request, id):
    """Editar un cliente existente"""
    cliente = get_object_or_404(Cliente, id=id)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        nomenclatura = request.POST.get('nomenclatura', '').strip().upper()
        
        # Validaciones
        if not nombre or not nomenclatura:
            messages.error(request, 'Todos los campos son obligatorios')
            return render(request, 'catalogos/cliente_form.html', {'cliente': cliente})
        
        if len(nomenclatura) > 5:
            messages.error(request, 'La nomenclatura no puede tener más de 5 caracteres')
            return render(request, 'catalogos/cliente_form.html', {'cliente': cliente})
        
        # Verificar si la nomenclatura ya existe (excluyendo el actual)
        if Cliente.objects.filter(nomenclatura=nomenclatura).exclude(id=id).exists():
            messages.error(request, f'La nomenclatura "{nomenclatura}" ya existe')
            return render(request, 'catalogos/cliente_form.html', {'cliente': cliente})
        
        try:
            cliente.nombre = nombre
            cliente.nomenclatura = nomenclatura
            cliente.activo = request.POST.get('activo', 'on') == 'on'
            cliente.save()
            
            messages.success(request, f'Cliente "{cliente.nombre}" actualizado exitosamente')
            return redirect('extractor:clientes_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar cliente: {str(e)}')
    
    return render(request, 'catalogos/cliente_form.html', {'cliente': cliente})

@login_required
def cliente_delete(request, id):
    """Eliminar un cliente"""
    cliente = get_object_or_404(Cliente, id=id)
    
    if request.method == 'POST':
        try:
            nombre = cliente.nombre
            cliente.delete()
            messages.success(request, f'Cliente "{nombre}" eliminado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar cliente: {str(e)}')
    
    return redirect('extractor:clientes_list')

@login_required
def tipos_servicio_list(request):
    try:
        tipos = TipoServicio.objects.filter(activo=True)
        
        # Ordenamiento - con validación EXTRA
        orden = request.GET.get('orden', 'id')
        
        # Solo permitir campos que existen en el modelo
        campos_validos = ['id', 'nombre', 'nomenclatura', 'activo', 'fecha_creacion']
        
        orden_final = 'id'  # Valor por defecto
        
        if orden:
            orden_limpio = orden.lstrip('-')
            if orden_limpio in campos_validos:
                # Si el campo es válido, mantener el prefijo de orden
                orden_final = orden
            else:
                orden_final = 'id'
        
        # Aplicar ordenamiento SOLO si es seguro
        tipos = tipos.order_by(orden_final)
        
        # NO imprimas el query SQL directamente en producción
        # print(f"Query SQL: {tipos.query}")  ← COMENTA ESTA LÍNEA
        
        context = {
            'tipos': tipos,
        }
        return render(request, 'catalogos/tipos_servicio_list.html', context)
        
    except Exception as e:
        # Manejo de error mejorado
        print(f"ERROR EN tipo_servicio_list: {str(e)}")
        # Devolver lista vacía en caso de error
        context = {
            'tipos': TipoServicio.objects.none(),
            'error': str(e)
        }
        return render(request, 'catalogos/tipos_servicio_list.html', context)
        
    except Exception as e:
        error_traceback = traceback.format_exc()
        print(f"ERROR EN tipo_servicio_list: {str(e)}")
        print(f"Traceback: {error_traceback}")
        
        return HttpResponse(f"""
            <h1>Error en tipo_servicio_list</h1>
            <p><strong>Error:</strong> {str(e)}</p>
            <h2>Traceback:</h2>
            <pre>{error_traceback}</pre>
            <h2>GET parameters:</h2>
            <pre>{dict(request.GET)}</pre>
        """)

@login_required
def tipo_servicio_create(request):
    """Crear un nuevo tipo de servicio"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        nomenclatura = request.POST.get('nomenclatura', '').strip().upper()
        
        # Validaciones
        if not nombre or not nomenclatura:
            messages.error(request, 'Todos los campos son obligatorios')
            return render(request, 'catalogos/tipo_servicio_form.html')
        
        if len(nomenclatura) > 10:
            messages.error(request, 'La nomenclatura no puede tener más de 10 caracteres')
            return render(request, 'catalogos/tipo_servicio_form.html')
        
        # Verificar si ya existe la nomenclatura
        if TipoServicio.objects.filter(nomenclatura=nomenclatura).exists():
            messages.error(request, f'La nomenclatura "{nomenclatura}" ya existe')
            return render(request, 'catalogos/tipo_servicio_form.html')
        
        try:
            tipo_servicio = TipoServicio.objects.create(
                nombre=nombre,
                nomenclatura=nomenclatura,
                activo=request.POST.get('activo', 'on') == 'on'
            )
            messages.success(request, f'Tipo de servicio "{tipo_servicio.nombre}" creado exitosamente')
            return redirect('tipos_servicio_list')
        except Exception as e:
            messages.error(request, f'Error al crear tipo de servicio: {str(e)}')
    
    return render(request, 'catalogos/tipo_servicio_form.html')

@login_required
def tipo_servicio_edit(request, id):
    """Editar un tipo de servicio existente"""
    tipo = get_object_or_404(TipoServicio, id=id)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        nomenclatura = request.POST.get('nomenclatura', '').strip().upper()
        
        # Validaciones
        if not nombre or not nomenclatura:
            messages.error(request, 'Todos los campos son obligatorios')
            return render(request, 'catalogos/tipo_servicio_form.html', {'tipo': tipo})
        
        if len(nomenclatura) > 10:
            messages.error(request, 'La nomenclatura no puede tener más de 10 caracteres')
            return render(request, 'catalogos/tipo_servicio_form.html', {'tipo': tipo})
        
        # Verificar si la nomenclatura ya existe (excluyendo el actual)
        if TipoServicio.objects.filter(nomenclatura=nomenclatura).exclude(id=id).exists():
            messages.error(request, f'La nomenclatura "{nomenclatura}" ya existe')
            return render(request, 'catalogos/tipo_servicio_form.html', {'tipo': tipo})
        
        try:
            tipo.nombre = nombre
            tipo.nomenclatura = nomenclatura
            tipo.activo = request.POST.get('activo', 'on') == 'on'
            tipo.save()
            
            messages.success(request, f'Tipo de servicio "{tipo.nombre}" actualizado exitosamente')
            return redirect('tipos_servicio_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar tipo de servicio: {str(e)}')
    
    return render(request, 'catalogos/tipo_servicio_form.html', {'tipo': tipo})


@login_required
def tipo_servicio_delete(request, id):
    """Eliminar un tipo de servicio"""
    tipo = get_object_or_404(TipoServicio, id=id)
    
    if request.method == 'POST':
        try:
            nombre = tipo.nombre
            tipo.delete()
            messages.success(request, f'Tipo de servicio "{nombre}" eliminado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar tipo de servicio: {str(e)}')
    
    return redirect('tipos_servicio_list')

@login_required
def proyectos_list(request):
    """Lista todos los proyectos con filtro por cliente opcional"""
    cliente_id = request.GET.get('cliente', '')
    
    if cliente_id:
        cliente = get_object_or_404(Cliente, id=cliente_id)
        proyectos = Proyecto.objects.filter(cliente=cliente).order_by('nombre')
    else:
        proyectos = Proyecto.objects.all().order_by('cliente__nombre', 'nombre')
    
    clientes = Cliente.objects.filter(activo=True).order_by('nombre')
    
    return render(request, 'catalogos/proyectos_list.html', {
        'proyectos': proyectos,
        'clientes': clientes,
        'cliente_filtro': cliente_id
    })
@login_required
def proyecto_create(request):
    """Crear un nuevo proyecto"""
    clientes = Cliente.objects.filter(activo=True).order_by('nombre')
    
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente', '')
        nombre = request.POST.get('nombre', '').strip()
        codigo = request.POST.get('codigo', '').strip().upper()
        
        # Validaciones
        if not cliente_id or not nombre or not codigo:
            messages.error(request, 'Todos los campos obligatorios deben completarse')
            return render(request, 'catalogos/proyecto_form.html', {'clientes': clientes})
        
        try:
            cliente = Cliente.objects.get(id=cliente_id)
        except Cliente.DoesNotExist:
            messages.error(request, 'Cliente no válido')
            return render(request, 'catalogos/proyecto_form.html', {'clientes': clientes})
        
        # Verificar si ya existe el código
        if Proyecto.objects.filter(codigo=codigo).exists():
            messages.error(request, f'El código "{codigo}" ya existe')
            return render(request, 'catalogos/proyecto_form.html', {'clientes': clientes})
        
        # Verificar si el cliente ya tiene un proyecto con el mismo nombre
        if Proyecto.objects.filter(cliente=cliente, nombre=nombre).exists():
            messages.error(request, f'Este cliente ya tiene un proyecto con el nombre "{nombre}"')
            return render(request, 'catalogos/proyecto_form.html', {'clientes': clientes})
        
        try:
            proyecto = Proyecto.objects.create(
                cliente=cliente,
                nombre=nombre,
                codigo=codigo,
                descripcion=request.POST.get('descripcion', '').strip(),
                activo=request.POST.get('activo', 'on') == 'on',
                fecha_inicio=request.POST.get('fecha_inicio') or None,
                fecha_fin=request.POST.get('fecha_fin') or None
            )
            messages.success(request, f'Proyecto "{proyecto.nombre}" creado exitosamente')
            return redirect('extractor:proyectos_list')
        except Exception as e:
            messages.error(request, f'Error al crear proyecto: {str(e)}')
    
    return render(request, 'catalogos/proyecto_form.html', {'clientes': clientes})

@login_required
def proyecto_edit(request, id):
    """Editar un proyecto existente"""
    proyecto = get_object_or_404(Proyecto, id=id)
    clientes = Cliente.objects.filter(activo=True).order_by('nombre')
    
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente', '')
        nombre = request.POST.get('nombre', '').strip()
        codigo = request.POST.get('codigo', '').strip().upper()
        
        # Validaciones
        if not cliente_id or not nombre or not codigo:
            messages.error(request, 'Todos los campos obligatorios deben completarse')
            return render(request, 'catalogos/proyecto_form.html', {
                'proyecto': proyecto,
                'clientes': clientes
            })
        
        try:
            cliente = Cliente.objects.get(id=cliente_id)
        except Cliente.DoesNotExist:
            messages.error(request, 'Cliente no válido')
            return render(request, 'catalogos/proyecto_form.html', {
                'proyecto': proyecto,
                'clientes': clientes
            })
        
        # Verificar si ya existe el código (excluyendo el actual)
        if Proyecto.objects.filter(codigo=codigo).exclude(id=id).exists():
            messages.error(request, f'El código "{codigo}" ya existe')
            return render(request, 'catalogos/proyecto_form.html', {
                'proyecto': proyecto,
                'clientes': clientes
            })
        
        # Verificar si el cliente ya tiene un proyecto con el mismo nombre (excluyendo el actual)
        if Proyecto.objects.filter(cliente=cliente, nombre=nombre).exclude(id=id).exists():
            messages.error(request, f'Este cliente ya tiene un proyecto con el nombre "{nombre}"')
            return render(request, 'catalogos/proyecto_form.html', {
                'proyecto': proyecto,
                'clientes': clientes
            })
        
        try:
            proyecto.cliente = cliente
            proyecto.nombre = nombre
            proyecto.codigo = codigo
            proyecto.descripcion = request.POST.get('descripcion', '').strip()
            proyecto.activo = request.POST.get('activo', 'on') == 'on'
            proyecto.fecha_inicio = request.POST.get('fecha_inicio') or None
            proyecto.fecha_fin = request.POST.get('fecha_fin') or None
            proyecto.save()
            
            messages.success(request, f'Proyecto "{proyecto.nombre}" actualizado exitosamente')
            return redirect('extractor:proyectos_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar proyecto: {str(e)}')
    
    return render(request, 'catalogos/proyecto_form.html', {
        'proyecto': proyecto,
        'clientes': clientes
    })

@login_required
def proyecto_delete(request, id):
    """Eliminar un proyecto"""
    proyecto = get_object_or_404(Proyecto, id=id)
    
    if request.method == 'POST':
        try:
            nombre = proyecto.nombre
            proyecto.delete()
            messages.success(request, f'Proyecto "{nombre}" eliminado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar proyecto: {str(e)}')
    
    return redirect('extractor:proyectos_list')

    
def generate_and_save_ticket(extracted_data, tipo_servicio_form, nomenclaturas, objetos_encontrados, request=None):
    """Genera y guarda un ticket en la base de datos"""
    
    # Obtener valores para los argumentos
    tipo_servicio_code = tipo_servicio_form
    tipo_pruebas_nom = nomenclaturas.get('tipo_pruebas_nomenclatura', '???')
    tipo_pruebas_id = objetos_encontrados.get('tipo_servicio_obj').id
    cliente_nom = nomenclaturas.get('cliente_nomenclatura', '???')
    proyecto_nom = nomenclaturas.get('proyecto_nomenclatura', '???')
    
    print(f"\n=== GENERANDO TICKET ===")
    print(f"Tipo Servicio: {tipo_servicio_code}")
    print(f"Tipo Pruebas NOMENCLATURA: {tipo_pruebas_nom}")
    print(f"Tipo Pruebas ID: {tipo_pruebas_id}")
    print(f"Cliente NOM: {cliente_nom}")
    print(f"Proyecto NOM: {proyecto_nom}")
    print("=======================\n")
    
    # Obtener el siguiente consecutivo
    consecutivo = get_next_consecutivo(
        tipo_servicio_code=tipo_servicio_code,
        tipo_pruebas_nom=tipo_pruebas_nom,
        tipo_pruebas_id=tipo_pruebas_id,
        cliente_nom=cliente_nom,
        proyecto_nom=proyecto_nom
    )
    
    # Convertir a entero y formatear
    consecutivo_num = int(consecutivo)
    consecutivo_str = f"{consecutivo_num:03d}"
    
    # Generar las partes del código
    empresa_code = "BID"
    
    ticket_code = f"{empresa_code}-{tipo_servicio_code}-{tipo_pruebas_nom}-{tipo_pruebas_id}-{cliente_nom}-{proyecto_nom}-{consecutivo_str}"
    
    print(f"🎫 Código de ticket generado: {ticket_code}")
    
    # Buscar los objetos relacionados
    cliente_obj = objetos_encontrados.get('cliente_obj')
    proyecto_obj = objetos_encontrados.get('proyecto_obj')
    tipo_servicio_obj = objetos_encontrados.get('tipo_servicio_obj')
    
    # Crear el ticket en la base de datos
    try:
        ticket_data = {
            'codigo': ticket_code,
            'empresa_code': empresa_code,
            'tipo_servicio_code': tipo_servicio_code,
            'funcion_code': tipo_pruebas_nom,
            'version_code': str(tipo_pruebas_id),
            'cliente_code': cliente_nom,
            'proyecto_code': proyecto_nom,
            'consecutivo': consecutivo_num,
            'cliente': cliente_obj,
            'proyecto': proyecto_obj,
            'tipo_servicio': tipo_servicio_obj,
            'responsable_solicitud': extracted_data.get('responsable_solicitud', ''),
            'lider_proyecto': extracted_data.get('lider_proyecto', ''),
            'numero_version': extracted_data.get('numero_version', ''),
            'estado': 'GENERADO',  # ← CAMBIAR DE 'ABIERTO' a 'GENERADO'
            'fecha_creacion': timezone.now(),  # ← AGREGAR EXPLÍCITAMENTE LA FECHA
        }
        
        # Agregar el usuario creador si request está disponible
        if request and request.user.is_authenticated:
            ticket_data['creado_por'] = request.user
            # Por defecto, asignar al mismo usuario
            ticket_data['asignado_a'] = request.user
        
        ticket = Ticket.objects.create(**ticket_data)
        
        # Registrar comentario inicial
        if request and request.user.is_authenticated:
            comentario_inicial = f"Ticket creado por {request.user.get_full_name() or request.user.username} vía carga de Excel"
            ticket.comentarios_seguimiento = comentario_inicial
            ticket.save()
        
        print(f"✅ Ticket guardado en BD con ID: {ticket.id}")
        return ticket_code, ticket
        
    except Exception as e:
        print(f"❌ Error al guardar ticket: {str(e)}")
        import traceback
        traceback.print_exc()
        return ticket_code, None

@login_required
def ticket_list(request):
    """Listado de tickets con filtros y paginación"""
    from django.db import connection
    from django.utils import timezone
    
     
    tickets = Ticket.objects.all().select_related('cliente', 'proyecto', 'tipo_servicio')

    # Filtros
    estado = request.GET.get('estado')
    cliente_id = request.GET.get('cliente')
    proyecto_id = request.GET.get('proyecto')
    busqueda = request.GET.get('q')
    por_pagina = request.GET.get('por_pagina', 20)

    if estado:
        tickets = tickets.filter(estado=estado)
    if cliente_id:
        tickets = tickets.filter(cliente_id=cliente_id)
    if proyecto_id:
        tickets = tickets.filter(proyecto_id=proyecto_id)
    if busqueda:
        tickets = tickets.filter(
            Q(codigo__icontains=busqueda) |
            Q(responsable_solicitud__icontains=busqueda) |
            Q(lider_proyecto__icontains=busqueda)
        )

    # Ordenamiento - los más nuevos primero por defecto
    orden = request.GET.get('orden', '-fecha_creacion')
    tickets = tickets.order_by(orden)

    # PAGINACIÓN
    try:
        por_pagina = int(por_pagina)
        if por_pagina not in [10, 20, 50, 100]:
            por_pagina = 20
    except ValueError:
        por_pagina = 20
    
    paginator = Paginator(tickets, por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Debug: imprimir información de tickets
    print(f"📊 Total tickets en BD: {Ticket.objects.count()}")
    print(f"📊 Tickets en página: {len(page_obj)}")
    for t in page_obj:
        print(f"   - {t.codigo} | Estado: {t.estado} | Fecha: {t.fecha_creacion}")

    # Estadísticas - CORREGIDO para incluir 'ABIERTO' si existe
    tickets_generados = Ticket.objects.filter(estado='GENERADO').count()
    tickets_abiertos = Ticket.objects.filter(estado='ABIERTO').count()  # ← NUEVO
    tickets_proceso = Ticket.objects.filter(estado='EN_PROCESO').count()
    tickets_completados = Ticket.objects.filter(estado='COMPLETADO').count()
    tickets_cancelados = Ticket.objects.filter(estado='CANCELADO').count()
    
    context = {
        'tickets': page_obj,
        'page_obj': page_obj,
        'total_tickets': Ticket.objects.count(),
        'tickets_generados': tickets_generados + tickets_abiertos,  # ← SUMAR AMBOS
        'tickets_proceso': tickets_proceso,
        'tickets_completados': tickets_completados,
        'tickets_cancelados': tickets_cancelados,
        'clientes': Cliente.objects.filter(activo=True),
        'tipos_servicio': TipoServicio.objects.filter(activo=True),
        'proyectos': Proyecto.objects.filter(activo=True).select_related('cliente'),
        'estados_disponibles': Ticket.ESTADOS_TICKET,
        'estado_selected': estado,
        'cliente_selected': int(cliente_id) if cliente_id else 0,
        'proyecto_selected': int(proyecto_id) if proyecto_id else 0,
        'busqueda': busqueda or '',
        'orden_actual': orden,
        'por_pagina': por_pagina,
        'tickets_count': tickets.count(),
    }
    return render(request, 'catalogos/ticket_list.html', context)

@login_required
def ticket_delete(request, id):
    """Eliminar un ticket"""
    ticket = get_object_or_404(Ticket, id=id)
    
    if request.method == 'POST':
        try:
            codigo = ticket.codigo
            
            # Si el ticket tiene datos Excel asociados, también se eliminarán (on_delete=CASCADE)
            ticket.delete()
            
            messages.success(request, f'✅ Ticket "{codigo}" eliminado exitosamente')
            return redirect('extractor:ticket_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar ticket: {str(e)}')
            return redirect('extractor:ticket_list')
    
    # GET request - mostrar página de confirmación
    context = {
        'ticket': ticket,
    }
    return render(request, 'catalogos/ticket_confirm_delete.html', context)

@login_required
def ticket_detail(request, id):  # Cambia 'ticket_id' por 'id'
    ticket = get_object_or_404(Ticket, id=id)
    
    # Preparar la lista de comentarios
    comentarios_lista = []
    if ticket.comentarios_seguimiento:
        comentarios_lista = ticket.comentarios_seguimiento.split('\n')
        comentarios_lista = [c for c in comentarios_lista if c.strip()]
    
    context = {
        'ticket': ticket,
        'comentarios_lista': comentarios_lista,
    }
    return render(request, 'catalogos/ticket_detail.html', context)

@login_required
def ticket_create(request):
    """Crear un nuevo ticket manualmente"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            cliente_id = request.POST.get('cliente')
            proyecto_id = request.POST.get('proyecto')
            tipo_prueba_id = request.POST.get('tipo_prueba')  # ID del TipoServicio (FUN, INT)
            tipo_servicio_code = request.POST.get('tipo_servicio_code')  # PRU, EST, G&A
            
            # Debug - imprime los valores recibidos
            print("\n=== DEBUG TICKET CREATE POST ===")
            print(f"cliente_id: '{cliente_id}'")
            print(f"proyecto_id: '{proyecto_id}'")
            print(f"tipo_prueba_id: '{tipo_prueba_id}'")
            print(f"tipo_servicio_code: '{tipo_servicio_code}'")
            print("================================\n")
            
            # VALIDACIÓN CORREGIDA - Verificar TODOS los campos obligatorios
            campos_faltantes = []
            
            if not cliente_id:
                campos_faltantes.append("Cliente")
            
            if not proyecto_id:
                campos_faltantes.append("Proyecto")
            
            if not tipo_prueba_id:
                campos_faltantes.append("Tipo de Prueba")
            
            if not tipo_servicio_code:
                campos_faltantes.append("Tipo de Servicio")
            
            # Si faltan campos, mostrar mensaje de error
            if campos_faltantes:
                mensaje = "Los siguientes campos son obligatorios: " + ", ".join(campos_faltantes)
                messages.error(request, mensaje)
                return redirect('extractor:ticket_create')
            
            # Obtener los objetos relacionados
            try:
                cliente = Cliente.objects.get(id=cliente_id, activo=True)
            except Cliente.DoesNotExist:
                messages.error(request, 'El cliente seleccionado no existe')
                return redirect('extractor:ticket_create')
            
            try:
                proyecto = Proyecto.objects.get(id=proyecto_id, activo=True)
            except Proyecto.DoesNotExist:
                messages.error(request, 'El proyecto seleccionado no existe')
                return redirect('extractor:ticket_create')
            
            try:
                tipo_prueba = TipoServicio.objects.get(id=tipo_prueba_id, activo=True)
            except TipoServicio.DoesNotExist:
                messages.error(request, 'El tipo de prueba seleccionado no existe')
                return redirect('extractor:ticket_create')
            
            # Verificar que el proyecto pertenezca al cliente
            if proyecto.cliente_id != cliente.id:
                messages.error(request, 'El proyecto seleccionado no pertenece al cliente')
                return redirect('extractor:ticket_create')
            
            # Validar consecutivo
            consecutivo_manual = request.POST.get('consecutivo', '').strip()
            
            if consecutivo_manual:
                try:
                    consecutivo_num = int(consecutivo_manual)
                    if consecutivo_num < 1 or consecutivo_num > 999:
                        messages.error(request, 'El consecutivo debe estar entre 1 y 999')
                        return redirect('extractor:ticket_create')
                    
                    # Verificar si ya existe
                    ticket_existente = Ticket.objects.filter(
                        empresa_code="BID",
                        tipo_servicio_code=tipo_servicio_code,
                        funcion_code=tipo_prueba.nomenclatura,
                        version_code=str(tipo_prueba.id),
                        cliente_code=cliente.nomenclatura,
                        proyecto_code=proyecto.codigo,
                        consecutivo=consecutivo_num
                    ).exists()
                    
                    if ticket_existente:
                        messages.error(request, f'Ya existe un ticket con el consecutivo {consecutivo_num:03d} para esta combinación')
                        return redirect('extractor:ticket_create')
                    
                except ValueError:
                    messages.error(request, 'El consecutivo debe ser un número válido')
                    return redirect('extractor:ticket_create')
            else:
                # Auto-generar consecutivo
                tickets_existentes = Ticket.objects.filter(
                    empresa_code="BID",
                    tipo_servicio_code=tipo_servicio_code,
                    funcion_code=tipo_prueba.nomenclatura,
                    version_code=str(tipo_prueba.id),
                    cliente_code=cliente.nomenclatura,
                    proyecto_code=proyecto.codigo
                )
                
                if tickets_existentes.exists():
                    max_consecutivo = tickets_existentes.aggregate(models.Max('consecutivo'))['consecutivo__max']
                    consecutivo_num = (max_consecutivo or 0) + 1
                else:
                    consecutivo_num = 1
            
            # Generar código del ticket
            consecutivo_str = f"{consecutivo_num:03d}"
            ticket_code = f"BID-{tipo_servicio_code}-{tipo_prueba.nomenclatura}-{tipo_prueba.id}-{cliente.nomenclatura}-{proyecto.codigo}-{consecutivo_str}"
            
            # Crear el ticket
            ticket = Ticket.objects.create(
                codigo=ticket_code,
                empresa_code="BID",
                tipo_servicio_code=tipo_servicio_code,
                funcion_code=tipo_prueba.nomenclatura,
                version_code=str(tipo_prueba.id),
                cliente_code=cliente.nomenclatura,
                proyecto_code=proyecto.codigo,
                consecutivo=consecutivo_num,
                cliente=cliente,
                proyecto=proyecto,
                tipo_servicio=tipo_prueba,
                responsable_solicitud=request.POST.get('responsable_solicitud', '')[:255],
                lider_proyecto=request.POST.get('lider_proyecto', '')[:255],
                numero_version=request.POST.get('numero_version', '')[:255],
                estado='GENERADO'
            )
            
            # Crear ExcelData si hay información adicional
            if any([
                request.POST.get('funcionalidad_liberacion'),
                request.POST.get('detalle_cambios'),
                request.POST.get('justificacion_cambio')
            ]):
                excel_data = ExcelData.objects.create(
                    cliente=str(cliente.id),
                    proyecto=str(proyecto.id),
                    tipo_pruebas=str(tipo_prueba.id),
                    tipo_servicio=tipo_servicio_code,
                    responsable_solicitud=request.POST.get('responsable_solicitud', ''),
                    lider_proyecto=request.POST.get('lider_proyecto', ''),
                    numero_version=request.POST.get('numero_version', ''),
                    funcionalidad_liberacion=request.POST.get('funcionalidad_liberacion', ''),
                    detalle_cambios=request.POST.get('detalle_cambios', ''),
                    justificacion_cambio=request.POST.get('justificacion_cambio', ''),
                    ticket_code=ticket_code
                )
                ticket.excel_data = excel_data
                ticket.save()
            
            messages.success(request, f'✅ Ticket creado exitosamente: {ticket_code}')
            return redirect('extractor:ticket_detail', id=ticket.id)
            
        except Exception as e:
            import traceback
            print(f"ERROR EN TICKET CREATE: {str(e)}")
            print(traceback.format_exc())
            messages.error(request, f'Error al crear ticket: {str(e)}')
            return redirect('extractor:ticket_create')
    
    # GET request - mostrar formulario
    clientes = Cliente.objects.filter(activo=True).order_by('nombre')
    tipos_servicio = TipoServicio.objects.filter(activo=True).order_by('nombre')
    
    # Obtener el último consecutivo
    ultimo_ticket = Ticket.objects.order_by('-consecutivo').first()
    ultimo_consecutivo = ultimo_ticket.consecutivo if ultimo_ticket else 0
    
    context = {
        'clientes': clientes,
        'tipos_servicio': tipos_servicio,
        'proyectos': [],
        'ultimo_consecutivo': ultimo_consecutivo,
    }
    return render(request, 'catalogos/new_ticket_form.html', context)


def proyectos_por_cliente(request, cliente_id):
    """Obtener proyectos de un cliente específico (para AJAX)"""
    try:
        # CORREGIDO: Usar el nombre correcto del campo
        proyectos = Proyecto.objects.filter(
            cliente_id=cliente_id, 
            activo=True
        ).order_by('nombre').values('id', 'nombre', 'codigo', 'nomenclatura')
        
        proyectos_list = list(proyectos)
        print(f"Proyectos encontrados para cliente {cliente_id}: {len(proyectos_list)}")
        
        return JsonResponse({'proyectos': proyectos_list})
        
    except Exception as e:
        print(f"Error en proyectos_por_cliente: {str(e)}")
        return JsonResponse({'error': str(e), 'proyectos': []})

@login_required
def ticket_create_simple(request):
    """VERSIÓN SIMPLIFICADA - Crear un nuevo ticket manualmente"""
    
    # GET - Mostrar formulario
    if request.method == 'GET':
        context = {
            'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
            'tipos_servicio': TipoServicio.objects.filter(activo=True).order_by('nombre'),
        }
        return render(request, 'catalogos/new_ticket_form_simple.html', context)
    
    # POST - Procesar formulario
    if request.method == 'POST':
        try:
            # 1. OBTENER DATOS BÁSICOS
            cliente_id = request.POST.get('cliente')
            proyecto_id = request.POST.get('proyecto')
            tipo_servicio_id = request.POST.get('tipo_servicio')
            
            # Validación MÍNIMA
            if not cliente_id or not proyecto_id or not tipo_servicio_id:
                messages.error(request, 'Cliente, Proyecto y Tipo de Servicio son obligatorios')
                return redirect('extractor:ticket_create_simple')
            
            # 2. OBTENER OBJETOS
            try:
                cliente = Cliente.objects.get(id=cliente_id, activo=True)
                proyecto = Proyecto.objects.get(id=proyecto_id, activo=True)
                tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
            except (Cliente.DoesNotExist, Proyecto.DoesNotExist, TipoServicio.DoesNotExist):
                messages.error(request, 'Uno de los elementos seleccionados no existe')
                return redirect('extractor:ticket_create_simple')
            
            # 3. VALIDAR QUE EL PROYECTO PERTENEZCA AL CLIENTE
            if proyecto.cliente_id != cliente.id:
                messages.error(request, 'El proyecto no pertenece al cliente seleccionado')
                return redirect('extractor:ticket_create_simple')
            
            # 4. PROCESAR CONSECUTIVO
            consecutivo_manual = request.POST.get('consecutivo', '').strip()
            
            if consecutivo_manual:
                # USAR CONSECUTIVO MANUAL
                try:
                    consecutivo_num = int(consecutivo_manual)
                    if consecutivo_num < 1 or consecutivo_num > 999:
                        messages.error(request, 'El consecutivo debe ser entre 1 y 999')
                        return redirect('extractor:ticket_create_simple')
                    
                    # Verificar si ya existe
                    existe = Ticket.objects.filter(
                        empresa_code="BID",
                        tipo_servicio_code=tipo_servicio.nomenclatura,
                        funcion_code=tipo_servicio.nomenclatura,
                        version_code=str(tipo_servicio.id),
                        cliente_code=cliente.nomenclatura,
                        proyecto_code=proyecto.codigo,
                        consecutivo=consecutivo_num
                    ).exists()
                    
                    if existe:
                        messages.error(request, f'Ya existe un ticket con consecutivo {consecutivo_num:03d}')
                        return redirect('extractor:ticket_create_simple')
                    
                    consecutivo_str = f"{consecutivo_num:03d}"
                    
                except ValueError:
                    messages.error(request, 'El consecutivo debe ser un número')
                    return redirect('extractor:ticket_create_simple')
            else:
                # AUTO-GENERAR CONSECUTIVO
                tickets_existentes = Ticket.objects.filter(
                    empresa_code="BID",
                    tipo_servicio_code=tipo_servicio.nomenclatura,
                    funcion_code=tipo_servicio.nomenclatura,
                    version_code=str(tipo_servicio.id),
                    cliente_code=cliente.nomenclatura,
                    proyecto_code=proyecto.codigo
                )
                
                if tickets_existentes.exists():
                    max_consecutivo = tickets_existentes.aggregate(models.Max('consecutivo'))['consecutivo__max']
                    consecutivo_num = (max_consecutivo or 0) + 1
                else:
                    consecutivo_num = 1
                
                consecutivo_str = f"{consecutivo_num:03d}"
            
            # 5. GENERAR CÓDIGO DEL TICKET
            ticket_code = f"BID-{tipo_servicio.nomenclatura}-{tipo_servicio.nomenclatura}-{tipo_servicio.id}-{cliente.nomenclatura}-{proyecto.codigo}-{consecutivo_str}"
            
            # 6. CREAR TICKET
            ticket = Ticket.objects.create(
                codigo=ticket_code,
                empresa_code="BID",
                tipo_servicio_code=tipo_servicio.nomenclatura,
                funcion_code=tipo_servicio.nomenclatura,
                version_code=str(tipo_servicio.id),
                cliente_code=cliente.nomenclatura,
                proyecto_code=proyecto.codigo,
                consecutivo=consecutivo_num,
                cliente=cliente,
                proyecto=proyecto,
                tipo_servicio=tipo_servicio,
                responsable_solicitud=request.POST.get('responsable_solicitud', '')[:255],
                lider_proyecto=request.POST.get('lider_proyecto', '')[:255],
                numero_version=request.POST.get('numero_version', '')[:255],
                estado='ABIERTO',  # Cambiado de GENERADO a ABIERTO
                creado_por=request.user,  # NUEVO
                asignado_a=request.user,  # NUEVO - Asignado al creador por defecto
                comentarios_seguimiento=f"Ticket creado manualmente por {request.user.get_full_name() or request.user.username}"
            )
            
            # 7. CREAR DATOS EXCEL ASOCIADOS (si hay información adicional)
            if any([
                request.POST.get('funcionalidad_liberacion'),
                request.POST.get('detalle_cambios'),
                request.POST.get('justificacion_cambio')
            ]):
                excel_data = ExcelData.objects.create(
                    cliente=str(cliente.id),
                    proyecto=str(proyecto.id),
                    tipo_pruebas=str(tipo_servicio.id),
                    tipo_servicio=tipo_servicio.nomenclatura,
                    responsable_solicitud=request.POST.get('responsable_solicitud', ''),
                    lider_proyecto=request.POST.get('lider_proyecto', ''),
                    numero_version=request.POST.get('numero_version', ''),
                    funcionalidad_liberacion=request.POST.get('funcionalidad_liberacion', ''),
                    detalle_cambios=request.POST.get('detalle_cambios', ''),
                    justificacion_cambio=request.POST.get('justificacion_cambio', ''),
                    ticket_code=ticket_code
                )
                ticket.excel_data = excel_data
                ticket.save()
            
            # 8. MENSAJE DE ÉXITO
            messages.success(request, f'✅ Ticket creado exitosamente: {ticket_code}')
            
            # Redirigir al detalle o al listado
            return redirect('extractor:ticket_detail', id=ticket.id)
            
        except Exception as e:
            # CAPTURAR CUALQUIER ERROR
            import traceback
            print(f"ERROR EN TICKET CREATE SIMPLE: {str(e)}")
            print(traceback.format_exc())
            
            messages.error(request, f'Error al crear ticket: {str(e)}')
            return redirect('extractor:ticket_create_simple')

@login_required
def generar_excel_dictamen(request, ticket_id):
    """
    Genera el Dictamen de Pruebas usando la plantilla data
    """
    import io
    import os
    from django.conf import settings
    from openpyxl import load_workbook
    from datetime import datetime, timedelta
    from django.contrib import messages
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Obtener la solicitud asociada (si existe)
    solicitud = None
    fecha_solicitud = None
    try:
        solicitud = SolicitudPruebas.objects.filter(ticket=ticket).first()
        if solicitud and solicitud.fecha_solicitud:
            fecha_solicitud = solicitud.fecha_solicitud
            print(f"✅ Fecha de solicitud encontrada: {fecha_solicitud}")
        else:
            print("⚠️ No se encontró solicitud asociada o no tiene fecha de solicitud")
            if ticket.fecha_creacion:
                fecha_solicitud = ticket.fecha_creacion
                print(f"   Usando fecha de creación del ticket: {fecha_solicitud}")
    except Exception as e:
        print(f"⚠️ Error al buscar solicitud: {e}")
    
    # Ruta a la plantilla
    plantilla_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'plantillas',
        'XXX-XXX-XXX-X-XXX-XXX-XXX DictamenPruebas PRUEBAS.xlsx'
    )
    
    if not os.path.exists(plantilla_path):
        messages.error(
            request, 
            f"No se encontró la plantilla. Por favor, coloca el archivo en: {plantilla_path}"
        )
        return redirect('extractor:ticket_detail', id=ticket.id)
    
    try:
        wb = load_workbook(plantilla_path)
        
        if 'Dictamen' in wb.sheetnames:
            ws = wb['Dictamen']
        else:
            ws = wb.active
        
        def set_cell_value(sheet, coordinate, value):
            """Escribe un valor en una celda, manejando correctamente celdas fusionadas"""
            try:
                for merged_range in sheet.merged_cells.ranges:
                    if coordinate in merged_range:
                        top_left = merged_range.start_cell.coordinate
                        sheet[top_left] = value
                        return
                sheet[coordinate] = value
            except Exception as e:
                print(f"⚠️ Error en {coordinate}: {e}")
        
        # Desglosar código del ticket
        partes = ticket.codigo.split('-')
        print(f"Partes del ticket: {partes}")
        
        if len(partes) >= 7:
            set_cell_value(ws, 'G2', partes[1])  # Tipo de Servicio
            set_cell_value(ws, 'I2', partes[2])  # Tipo de pruebas
            set_cell_value(ws, 'K2', partes[3])  # No. Pruebas
            set_cell_value(ws, 'M2', partes[4])  # Cliente
            set_cell_value(ws, 'O2', partes[5])  # Nomenclatura del Proyecto
            set_cell_value(ws, 'Q2', partes[6])  # Consecutivo
            print(f"✅ Código del ticket mapeado correctamente")
        
        # Fecha actual
        fecha_actual = datetime.now()
        fecha_actual_str = fecha_actual.strftime('%d/%m/%Y')
        
        # ===== PERÍODO DE PRUEBAS EN K5 =====
        if fecha_solicitud:
            if hasattr(fecha_solicitud, 'strftime'):
                fecha_solicitud_str = fecha_solicitud.strftime('%d/%m/%Y')
            else:
                try:
                    fecha_obj = datetime.strptime(str(fecha_solicitud), '%Y-%m-%d')
                    fecha_solicitud_str = fecha_obj.strftime('%d/%m/%Y')
                except:
                    fecha_solicitud_str = str(fecha_solicitud)
        else:
            if ticket.fecha_creacion:
                fecha_solicitud_str = ticket.fecha_creacion.strftime('%d/%m/%Y')
            else:
                fecha_solicitud_str = fecha_actual_str
        
        periodo_pruebas = f"{fecha_solicitud_str} - {fecha_actual_str}"
        set_cell_value(ws, 'K5', periodo_pruebas)
        print(f"✅ K5 = {periodo_pruebas}")
        
        # ===== FECHA ACTUAL EN K6 =====
        set_cell_value(ws, 'K6', fecha_actual_str)
        print(f"✅ K6 = {fecha_actual_str}")
        
        # ===== CÁLCULO DE HORAS EN M18 =====
        horas_totales = 0
        
        if fecha_solicitud:
            if hasattr(fecha_solicitud, 'date'):
                fecha_solicitud_date = fecha_solicitud.date()
            else:
                fecha_solicitud_date = fecha_solicitud
            
            fecha_actual_date = fecha_actual.date()
            dias_habiles = calcular_dias_habiles(fecha_solicitud_date, fecha_actual_date)
            horas_totales = dias_habiles * 8
            print(f"📊 Días hábiles: {dias_habiles}, Horas: {horas_totales}")
        
        try:
            is_merged_m18 = False
            for merged_range in ws.merged_cells.ranges:
                if 'M18' in merged_range or ws['M18'].coordinate in merged_range:
                    is_merged_m18 = True
                    top_left_cell = ws[merged_range.start_cell.coordinate]
                    top_left_cell.value = horas_totales
                    top_left_cell.number_format = '0'
                    print(f"✅ M18 = {horas_totales}")
                    break
            
            if not is_merged_m18:
                ws['M18'] = horas_totales
                ws['M18'].number_format = '0'
                print(f"✅ M18 = {horas_totales}")
        except Exception as e:
            print(f"❌ Error al asignar M18: {e}")
        
        # ===== SOLO CAMPOS QUE DEBEN MODIFICARSE =====
        # NOTA: E6, B24, H24 NO se modifican para mantener sus textos originales
        campos = [
            ('B5', ticket.cliente.nombre if ticket.cliente else ''),
            ('B6', ticket.proyecto.nombre if ticket.proyecto else ''),
            ('C7', ticket.tipo_servicio.nombre if ticket.tipo_servicio else ''),
        ]
        
        for celda, valor in campos:
            set_cell_value(ws, celda, valor)
            print(f"✅ {celda} = {valor}")
        
        # ===== RESPONSABLE DEL TICKET EN F28 =====
        responsable_nombre = ""
        if ticket.asignado_a:
            responsable_nombre = ticket.asignado_a.get_full_name() or ticket.asignado_a.username
        elif ticket.creado_por:
            responsable_nombre = ticket.creado_por.get_full_name() or ticket.creado_por.username
        else:
            responsable_nombre = "No asignado"
        
        set_cell_value(ws, 'F28', responsable_nombre)
        print(f"✅ F28 = {responsable_nombre}")
        
        # Ajustar altura de fila 37
        try:
            ws.row_dimensions[37].height = 32.6
        except:
            pass
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{ticket.codigo} Dictamen Pruebas.xlsx"'
        
        return response
        
    except Exception as e:
        print(f"❌ Error al generar dictamen: {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"Error al generar dictamen: {str(e)}")
        return redirect('extractor:ticket_detail', id=ticket.id)

def verificar_plantilla(request):
    import os
    from django.conf import settings
    from django.http import HttpResponse
    
    # Posibles rutas a verificar
    rutas = [
        os.path.join(settings.BASE_DIR, 'static', 'plantillas', 'XXX-XXX-XXX-X-XXX-XXX-XXX DictamenPruebas PRUEBAS.xlsx'),
        os.path.join(settings.BASE_DIR, 'extractor', 'static', 'plantillas', 'XXX-XXX-XXX-X-XXX-XXX-XXX DictamenPruebas PRUEBAS.xlsx'),
        os.path.join(settings.MEDIA_ROOT, 'plantillas', 'XXX-XXX-XXX-X-XXX-XXX-XXX DictamenPruebas PRUEBAS.xlsx'),
    ]
    
    resultado = "<h1>Verificación de Plantilla</h1>"
    resultado += f"<p>BASE_DIR: {settings.BASE_DIR}</p>"
    
    for ruta in rutas:
        existe = os.path.exists(ruta)
        resultado += f"<p>Ruta: {ruta}<br>Existe: {existe}</p>"
        if existe:
            resultado += f"<p>✅ ¡ENCONTRADA AQUÍ!</p>"
    
    return HttpResponse(resultado)


def generar_excel_resultados(request, ticket_id):
    import io
    import os
    from datetime import datetime
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from django.conf import settings
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Ruta a la plantilla de resultados
    plantilla_resultados_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'plantillas',
        'XXX-XXX-XXX-X-XXX-XXX-XXX Documentación de Resultados.xlsx'
    )
    
    # Verificar si existe la plantilla
    if os.path.exists(plantilla_resultados_path):
        # Usar la plantilla
        wb = load_workbook(plantilla_resultados_path)
        ws = wb.active
        ws.title = "Resultados Pruebas"
    else:
        # Crear un nuevo workbook si no existe la plantilla
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados Pruebas"
        
        # Configurar anchos de columna básicos
        ws.column_dimensions['A'].width = 50  # Ancho para el detalle de cambios
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 30
    
    # Definir estilos (solo si creamos el documento desde cero)
    if not os.path.exists(plantilla_resultados_path):
        header_font = Font(bold=True)
        ticket_font = Font(bold=True, size=14, color="2563EB")
        
        # TICKET - SOLO SI ES DOCUMENTO NUEVO
        ws.cell(row=2, column=1, value="TICKET:")
        ws.cell(row=2, column=1).font = header_font
    
    # AGREGAR EL TICKET EN C2
    ws['C2'] = ticket.codigo
    
    # Si es documento nuevo, aplicar estilo al ticket
    if not os.path.exists(plantilla_resultados_path):
        ws['C2'].font = Font(bold=True, size=14, color="2563EB")
    
    # Versión (si no existe en la plantilla)
    if ws['M3'].value is None or "Versión" not in str(ws['M3'].value):
        ws['M3'] = f"VERSIÓN: Versión {ticket.numero_version or '1.0.0'}"
    
    # ===== NUEVO: COPIAR DETALLE DE CAMBIOS EN A8 =====
    if ticket.excel_data and ticket.excel_data.detalle_cambios:
        # Obtener el detalle de cambios
        detalle_cambios = ticket.excel_data.detalle_cambios
        
        # Limpiar y preparar el texto
        detalle_cambios = detalle_cambios.strip()
        
        # Escribir en A8
        ws['A8'] = detalle_cambios
        
        # Aplicar formato para mejor legibilidad
        ws['A8'].alignment = Alignment(
            wrap_text=True,
            vertical='top',
            horizontal='left'
        )
        
        # Ajustar altura de fila según contenido (aproximado)
        lineas = detalle_cambios.count('\n') + 1
        altura_estimada = min(lineas * 15, 300)  # Máximo 300 puntos
        ws.row_dimensions[8].height = altura_estimada
        
        print(f"✅ Detalle de cambios copiado a A8: {len(detalle_cambios)} caracteres, {lineas} líneas")
    else:
        print("⚠️ No hay detalle de cambios para copiar")
        
        # Si no hay detalle de cambios, poner un placeholder
        ws['A8'] = "No se especificaron detalles de cambios."
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Crear respuesta con el nombre del archivo basado en el ticket
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"{ticket.codigo} Documentación de Resultados.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = len(buffer.getvalue())
    
    return response


def calcular_dias_habiles(fecha_inicio, fecha_fin):
    """
    Calcula el número de días hábiles entre dos fechas (excluyendo sábados y domingos)
    """
    from datetime import timedelta
    
    if not fecha_inicio or not fecha_fin:
        return 0
    
    # Asegurar que las fechas son datetime.date
    if hasattr(fecha_inicio, 'date'):
        fecha_inicio = fecha_inicio.date()
    if hasattr(fecha_fin, 'date'):
        fecha_fin = fecha_fin.date()
    
    # Si la fecha de inicio es mayor que la fecha de fin, retornar 0
    if fecha_inicio > fecha_fin:
        return 0
    
    # Contar días hábiles
    dias_habiles = 0
    fecha_actual = fecha_inicio
    
    while fecha_actual <= fecha_fin:
        # 0 = lunes, 1 = martes, ..., 4 = viernes, 5 = sábado, 6 = domingo
        if fecha_actual.weekday() < 5:  # Lunes a viernes
            dias_habiles += 1
        fecha_actual += timedelta(days=1)
    
    return dias_habiles


@login_required
def export_tickets_excel(request):
    """
    Exporta los tickets filtrados a un archivo Excel
    """
    # Obtener los mismos filtros que en ticket_list
    tickets = Ticket.objects.all().select_related('cliente', 'proyecto', 'tipo_servicio', 'excel_data')
    
    # Aplicar los mismos filtros que en la vista list
    estado = request.GET.get('estado')
    cliente_id = request.GET.get('cliente')
    proyecto_id = request.GET.get('proyecto')
    busqueda = request.GET.get('q')
    
    if estado:
        tickets = tickets.filter(estado=estado)
    if cliente_id:
        tickets = tickets.filter(cliente_id=cliente_id)
    if proyecto_id:
        tickets = tickets.filter(proyecto_id=proyecto_id)
    if busqueda:
        tickets = tickets.filter(
            Q(codigo__icontains=busqueda) |
            Q(responsable_solicitud__icontains=busqueda) |
            Q(lider_proyecto__icontains=busqueda)
        )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    # Definir estilos
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    
    # Definir encabezados
    headers = [
        'ID', 'Código Ticket', 'Estado', 'Cliente', 'Proyecto', 
        'Tipo Servicio', 'Responsable Solicitud', 'Líder Proyecto',
        'Versión', 'Funcionalidad', 'Detalle Cambios', 'Justificación',
        'Fecha Creación', 'Fecha Actualización'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos
    for row, ticket in enumerate(tickets, 2):
        ws.cell(row=row, column=1, value=ticket.id)
        ws.cell(row=row, column=2, value=ticket.codigo)
        ws.cell(row=row, column=3, value=ticket.get_estado_display())
        ws.cell(row=row, column=4, value=ticket.cliente.nombre if ticket.cliente else '')
        ws.cell(row=row, column=5, value=ticket.proyecto.nombre if ticket.proyecto else '')
        ws.cell(row=row, column=6, value=ticket.tipo_servicio.nombre if ticket.tipo_servicio else '')
        ws.cell(row=row, column=7, value=ticket.responsable_solicitud)
        ws.cell(row=row, column=8, value=ticket.lider_proyecto)
        ws.cell(row=row, column=9, value=ticket.numero_version)
        
        # Datos del Excel asociado
        excel_data = ticket.excel_data
        ws.cell(row=row, column=10, value=excel_data.funcionalidad_liberacion if excel_data else '')
        ws.cell(row=row, column=11, value=excel_data.detalle_cambios if excel_data else '')
        ws.cell(row=row, column=12, value=excel_data.justificacion_cambio if excel_data else '')
        
        ws.cell(row=row, column=13, value=ticket.fecha_creacion.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=14, value=ticket.fecha_actualizacion.strftime('%d/%m/%Y %H:%M'))
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Crear respuesta
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"tickets_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def ticket_cambiar_estado(request, id):
    """API para cambiar el estado de un ticket"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nuevo_estado = data.get('estado')
            
            ticket = get_object_or_404(Ticket, id=id)
            
            estados_validos = ['ABIERTO', 'GENERADO', 'EN_PROCESO', 'COMPLETADO', 'CANCELADO']
            if nuevo_estado not in estados_validos:
                return JsonResponse({'success': False, 'error': 'Estado no válido'})
            
            usuario = request.user.get_full_name() or request.user.username
            
            # ✅ CORRECCIÓN: Convertir a hora local
            from django.utils import timezone
            ahora_local = timezone.localtime(timezone.now())
            fecha_hora = ahora_local.strftime('%d/%m/%Y %H:%M')
            
            comentario_estado = f"[{fecha_hora}] {usuario} cambió el estado de {ticket.get_estado_display()} a {dict(Ticket.ESTADOS_TICKET).get(nuevo_estado)}"
            
            if ticket.comentarios_seguimiento:
                ticket.comentarios_seguimiento += f"\n{comentario_estado}"
            else:
                ticket.comentarios_seguimiento = comentario_estado
            
            if nuevo_estado == 'COMPLETADO' and ticket.estado != 'COMPLETADO':
                ticket.fecha_cierre = timezone.now()
            elif nuevo_estado != 'COMPLETADO':
                ticket.fecha_cierre = None
            
            ticket.estado = nuevo_estado
            ticket.save()
            
            estado_display = dict(Ticket.ESTADOS_TICKET).get(nuevo_estado)
            
            return JsonResponse({
                'success': True,
                'estado_display': estado_display
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def ticket_agregar_comentario(request, id):
    """API para agregar comentario de seguimiento"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            comentario = data.get('comentario', '').strip()
            
            if not comentario:
                return JsonResponse({'success': False, 'error': 'Comentario vacío'})
            
            ticket = get_object_or_404(Ticket, id=id)
            
            usuario = request.user.get_full_name() or request.user.username
            
            # ✅ CORRECCIÓN: Convertir a hora local
            from django.utils import timezone
            ahora_local = timezone.localtime(timezone.now())
            fecha_hora = ahora_local.strftime('%d/%m/%Y %H:%M')
            
            comentario_formateado = f"[{fecha_hora}] {usuario}: {comentario}"
            
            if ticket.comentarios_seguimiento:
                ticket.comentarios_seguimiento += f"\n{comentario_formateado}"
            else:
                ticket.comentarios_seguimiento = comentario_formateado
            
            ticket.save()
            
            return JsonResponse({
                'success': True,
                'comentario_formateado': comentario_formateado
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def export_table_csv(request, table_name):
    """
    Exporta una tabla específica a formato CSV
    """
    try:
        # Mapeo de nombres de tabla a modelos
        models_map = {
            'cliente': Cliente,
            'proyecto': Proyecto,
            'tiposervicio': TipoServicio,
            'ticket': Ticket,
            'exceldata': ExcelData,
            'solicitudpruebas': SolicitudPruebas,
        }
        
        if table_name.lower() not in models_map:
            return HttpResponse("Tabla no encontrada", status=404)
        
        model = models_map[table_name.lower()]
        queryset = model.objects.all()
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        # Agregar BOM para UTF-8 para que Excel lo abra correctamente
        response.write('\ufeff'.encode('utf-8'))  # BOM para UTF-8
        
        filename = f"{table_name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Obtener nombres de campos
        headers = [field.name for field in model._meta.fields]
        writer.writerow(headers)
        
        # Escribir datos
        for obj in queryset:
            row = []
            for field in headers:
                value = getattr(obj, field)
                # Manejar fechas y relaciones
                if value is None:
                    row.append('')
                elif hasattr(value, 'strftime'):  # Es una fecha
                    row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                elif hasattr(value, 'pk'):  # Es una relación
                    row.append(value.pk)
                else:
                    row.append(str(value))
            writer.writerow(row)
        
        return response
        
    except Exception as e:
        print(f"ERROR en export_table_csv: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponseServerError(f"Error al exportar: {str(e)}")

@login_required
def crear_ticket_manual(request):
    """Vista para crear solicitud de pruebas manualmente"""
    from django.utils import timezone
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            cliente_id = request.POST.get('cliente')
            proyecto_id = request.POST.get('proyecto')
            tipo_servicio_code = request.POST.get('tipo_servicio_code')  # PRU, EST, G&A
            tipo_prueba_id = request.POST.get('tipo_prueba')
            
            # Validaciones básicas
            if not cliente_id or not proyecto_id or not tipo_servicio_code or not tipo_prueba_id:
                messages.error(request, 'Todos los campos obligatorios deben estar llenos')
                return redirect('extractor:crear_ticket_manual')
            
            # Obtener objetos
            cliente = Cliente.objects.get(id=cliente_id, activo=True)
            proyecto = Proyecto.objects.get(id=proyecto_id, activo=True)
            tipo_prueba = TipoServicio.objects.get(id=tipo_prueba_id, activo=True)
            
            # Validar que el proyecto pertenezca al cliente
            if proyecto.cliente_id != cliente.id:
                messages.error(request, 'El proyecto no pertenece al cliente seleccionado')
                return redirect('extractor:crear_ticket_manual')
            
            # Generar consecutivo
            tickets_existentes = Ticket.objects.filter(
                empresa_code="BID",
                tipo_servicio_code=tipo_servicio_code,
                funcion_code=tipo_prueba.nomenclatura,
                version_code=str(tipo_prueba.id),
                cliente_code=cliente.nomenclatura,
                proyecto_code=proyecto.codigo
            )
            
            if tickets_existentes.exists():
                max_consecutivo = tickets_existentes.aggregate(models.Max('consecutivo'))['consecutivo__max']
                consecutivo_num = (max_consecutivo or 0) + 1
            else:
                consecutivo_num = 1
            
            consecutivo_str = f"{consecutivo_num:03d}"
            
            # Generar código del ticket
            ticket_code = f"BID-{tipo_servicio_code}-{tipo_prueba.nomenclatura}-{tipo_prueba.id}-{cliente.nomenclatura}-{proyecto.codigo}-{consecutivo_str}"
            
            # Crear el ticket
            ticket = Ticket.objects.create(
                codigo=ticket_code,
                empresa_code="BID",
                tipo_servicio_code=tipo_servicio_code,
                funcion_code=tipo_prueba.nomenclatura,
                version_code=str(tipo_prueba.id),
                cliente_code=cliente.nomenclatura,
                proyecto_code=proyecto.codigo,
                consecutivo=consecutivo_num,
                cliente=cliente,
                proyecto=proyecto,
                tipo_servicio=tipo_prueba,
                responsable_solicitud=request.POST.get('responsable_solicitud', '')[:255],
                lider_proyecto=request.POST.get('lider_proyecto', '')[:255],
                numero_version=request.POST.get('numero_version', '')[:255],
                estado='GENERADO'
            )
            
            # Crear ExcelData asociado
            excel_data = ExcelData.objects.create(
                cliente=str(cliente.id),
                proyecto=str(proyecto.id),
                tipo_pruebas=str(tipo_prueba.id),
                tipo_servicio=tipo_servicio_code,
                responsable_solicitud=request.POST.get('responsable_solicitud', ''),
                lider_proyecto=request.POST.get('lider_proyecto', ''),
                tipo_aplicacion=request.POST.get('tipo_aplicacion', ''),
                numero_version=request.POST.get('numero_version', ''),
                funcionalidad_liberacion=request.POST.get('funcionalidad_liberacion', ''),
                detalle_cambios=request.POST.get('detalle_cambios', ''),
                justificacion_cambio=request.POST.get('justificacion_cambio', ''),
                ticket_code=ticket_code
            )
            
            ticket.excel_data = excel_data
            ticket.save()
            
            messages.success(request, f'✅ Solicitud creada exitosamente. Ticket: {ticket_code}')
            return redirect('extractor:ticket_detail', id=ticket.id)
            
        except Exception as e:
            import traceback
            print(f"ERROR: {str(e)}")
            print(traceback.format_exc())
            messages.error(request, f'Error al crear solicitud: {str(e)}')
            return redirect('extractor:crear_ticket_manual')
    
    # GET - Mostrar formulario
    context = {
        'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
        'tipos_servicio': TipoServicio.objects.filter(activo=True).order_by('nombre'),
        'today': timezone.now().date(),
        'now': timezone.now(),
    }
    return render(request, 'extractor/crear_solicitud.html', context)

@login_required
def export_all_tables_backup(request):
    """
    Exporta todas las tablas como CSV en un archivo ZIP
    """
    try:
        # Crear archivo ZIP en memoria
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            models_to_export = {
                'clientes': Cliente,
                'proyectos': Proyecto,
                'tipos_servicio': TipoServicio,
                'tickets': Ticket,
                'datos_excel': ExcelData,
                'solicitudes_pruebas': SolicitudPruebas,
            }
            
            for filename, model in models_to_export.items():
                # Crear CSV en memoria usando StringIO para texto
                import io
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                
                queryset = model.objects.all()
                
                # Escribir encabezados
                headers = [field.name for field in model._meta.fields]
                writer.writerow(headers)
                
                # Escribir datos
                for obj in queryset:
                    row = []
                    for field in headers:
                        value = getattr(obj, field)
                        if value is None:
                            row.append('')
                        elif hasattr(value, 'strftime'):
                            row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                        elif hasattr(value, 'pk'):
                            row.append(value.pk)
                        else:
                            row.append(str(value))
                    writer.writerow(row)
                
                # Convertir StringIO a bytes para el ZIP
                csv_content = csv_buffer.getvalue().encode('utf-8-sig')  # UTF-8 con BOM para Excel
                zip_file.writestr(f"{filename}.csv", csv_content)
        
        # Preparar respuesta
        zip_buffer.seek(0)
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="backup_completo_{timestamp}.zip"'
        
        return response
        
    except Exception as e:
        print(f"ERROR en export_all_tables_backup: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponseServerError(f"Error al crear backup: {str(e)}")
    

def crear_solicitud(request):
    """
    Vista para crear solicitud de pruebas manualmente
    Con protecciones anti-bots (Cooldown + Honeypot + Rate Limiting)
    """
    from django.conf import settings
    
    # ===== CAPA 1: COOLDOWN DE 5 MINUTOS =====
    ultima_solicitud = request.session.get('ultima_solicitud_timestamp')
    cooldown_segundos = getattr(settings, 'SOLICITUD_COOLDOWN_SEGUNDOS', 300)  # 5 minutos por defecto
    
    tiempo_restante = 0
    if ultima_solicitud:
        tiempo_actual = timezone.now().timestamp()
        tiempo_transcurrido = tiempo_actual - ultima_solicitud
        
        if tiempo_transcurrido < cooldown_segundos:
            tiempo_restante = cooldown_segundos - tiempo_transcurrido
            
            if request.method == 'GET':
                minutos = int(tiempo_restante // 60)
                segundos = int(tiempo_restante % 60)
                messages.info(
                    request,
                    f'⏳ Puedes crear una nueva solicitud en {minutos} minutos y {segundos} segundos.'
                )
            
            if request.method == 'POST':
                minutos = int(tiempo_restante // 60)
                segundos = int(tiempo_restante % 60)
                messages.warning(
                    request, 
                    f'⏳ Debes esperar {minutos} minutos y {segundos} segundos antes de crear otra solicitud.'
                )
                return redirect('extractor:solicitud_list')
    
    # ===== CAPA 2: HONEYPOT (solo en POST) =====
    if request.method == 'POST':
        # Verificar campos honeypot (deben estar vacíos)
        if request.POST.get('web_contacto', ''):  # Si el campo oculto tiene contenido
            messages.error(request, 'Actividad sospechosa detectada. Si eres humano, no llenes campos ocultos.')
            logger.warning(f"Intento de bot detectado - IP: {request.META.get('REMOTE_ADDR')} - Usuario: {request.user}")
            return redirect('extractor:crear_solicitud')
        
        if request.POST.get('confirmar_email', ''):  # Segundo campo honeypot
            messages.error(request, 'Actividad sospechosa detectada.')
            logger.warning(f"Intento de bot detectado (campo2) - IP: {request.META.get('REMOTE_ADDR')}")
            return redirect('extractor:crear_solicitud')
    
    # ===== CAPA 3: RATE LIMITING POR IP (solo en POST) =====
    if request.method == 'POST':
        permitido, mensaje = check_rate_limit_by_ip(request, limite=5, tiempo_ventana=3600)
        if not permitido:
            messages.error(request, mensaje)
            logger.info(f"Rate limit excedido - IP: {request.META.get('REMOTE_ADDR')}")
            return redirect('extractor:solicitud_list')
    
    # ===== PROCESAR FORMULARIO (solo si pasa todas las capas) =====
    if request.method == 'POST':
        try:
            # ===== VALIDACIONES =====
            cliente_id = request.POST.get('cliente')
            proyecto_id = request.POST.get('proyecto')
            tipo_servicio_code = request.POST.get('tipo_servicio_code')
            tipo_prueba_id = request.POST.get('tipo_prueba')
            
            # Validaciones básicas
            if not cliente_id or not proyecto_id or not tipo_servicio_code or not tipo_prueba_id:
                messages.error(request, 'Los campos obligatorios deben estar llenos')
                return redirect('extractor:crear_solicitud')
            
            # Obtener objetos
            cliente = Cliente.objects.get(id=cliente_id, activo=True)
            proyecto = Proyecto.objects.get(id=proyecto_id, activo=True)
            tipo_prueba = TipoServicio.objects.get(id=tipo_prueba_id, activo=True)
            
            # Validar que el proyecto pertenezca al cliente
            if proyecto.cliente_id != cliente.id:
                messages.error(request, 'El proyecto no pertenece al cliente seleccionado')
                return redirect('extractor:crear_solicitud')
            
            # ===== CREAR SOLICITUD =====
            solicitud = SolicitudPruebas(
                cliente=cliente,
                proyecto=proyecto,
                fecha_solicitud=request.POST.get('fecha_solicitud') or timezone.now().date(),
                hora_solicitud=request.POST.get('hora_solicitud') or timezone.now().time(),
                tipo_servicio_code=tipo_servicio_code,
                tipo_prueba=tipo_prueba,
                area_solicitante=request.POST.get('area_solicitante', ''),
                numero_version=request.POST.get('numero_version', ''),
                responsable_solicitud=request.POST.get('responsable_solicitud', ''),
                lider_proyecto=request.POST.get('lider_proyecto', ''),
                tipo_aplicacion=request.POST.get('tipo_aplicacion', ''),
                funcionalidad_liberacion=request.POST.get('funcionalidad_liberacion', ''),
                detalle_cambios=request.POST.get('detalle_cambios', ''),
                justificacion_cambio=request.POST.get('justificacion_cambio', ''),
                puntos_considerar=request.POST.get('puntos_considerar', ''),
                pendientes=request.POST.get('pendientes', ''),
                insumos=request.POST.get('insumos', ''),
                creado_por=request.user.username if request.user.is_authenticated else 'Anónimo'
            )
            
            # Generar nombre antes de guardar
            solicitud.nombre_archivo = solicitud.generar_nombre_archivo()
            
            # Guardar solicitud
            solicitud.save()
            
            # ===== GUARDAR TIMESTAMP EN SESIÓN PARA COOLDOWN =====
            request.session['ultima_solicitud_timestamp'] = timezone.now().timestamp()
            request.session['ultima_solicitud_id'] = solicitud.id
            
            # <<< MODIFICACIÓN: En lugar de redirigir, mostrar panel de éxito >>>
            # Preparar el contexto para mostrar el panel de éxito
            context_exito = {
                'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
                'tipos_servicio': TipoServicio.objects.filter(activo=True).order_by('nombre'),
                'today': timezone.now().date(),
                'now': timezone.now(),
                'tiempo_restante': int(tiempo_restante) if 'tiempo_restante' in locals() else 0,
                'solicitud_creada': solicitud,  # Pasamos la solicitud creada
                'mostrar_resumen': True,  # Bandera para la plantilla
            }
            
            # Si se solicitó generar ticket ahora, lo generamos y lo añadimos al contexto
            if request.POST.get('generar_ticket_ahora') == 'on':
                ticket = solicitud.generar_ticket()
                context_exito['ticket_generado'] = ticket
                # Nota: ya no redirigimos a ticket_detail, solo pasamos el ticket al contexto
            
            # Renderizar la misma plantilla con el panel de éxito
            return render(request, 'extractor/crear_solicitud.html', context_exito)
            # >>> FIN MODIFICACIÓN
            
        except Cliente.DoesNotExist:
            messages.error(request, 'El cliente seleccionado no existe')
        except Proyecto.DoesNotExist:
            messages.error(request, 'El proyecto seleccionado no existe')
        except TipoServicio.DoesNotExist:
            messages.error(request, 'El tipo de prueba seleccionado no existe')
        except Exception as e:
            import traceback
            print(f"❌ Error al crear solicitud: {str(e)}")
            print(traceback.format_exc())
            messages.error(request, 'Error al crear solicitud. Por favor intenta de nuevo.')
        
        return redirect('extractor:crear_solicitud')
    
    # GET - Mostrar formulario
    context = {
        'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
        'tipos_servicio': TipoServicio.objects.filter(activo=True).order_by('nombre'),
        'today': timezone.now().date(),
        'now': timezone.now(),
        'tiempo_restante': int(tiempo_restante),  # Para el temporizador
    }
    return render(request, 'extractor/crear_solicitud.html', context)

@login_required
def solicitud_list(request):
    """Listado de solicitudes de pruebas"""
    from django.utils import timezone
    from datetime import timedelta

    today = timezone.now().date()
    week_ago = today - timedelta(days=7)

    solicitudes = SolicitudPruebas.objects.all().select_related('cliente', 'proyecto', 'tipo_prueba', 'ticket')
    # Filtros
    cliente_id = request.GET.get('cliente')
    proyecto_id = request.GET.get('proyecto')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    con_ticket = request.GET.get('con_ticket')
    sin_ticket = request.GET.get('sin_ticket')  
    
    if cliente_id:
        solicitudes = solicitudes.filter(cliente_id=cliente_id)
    if proyecto_id:
        solicitudes = solicitudes.filter(proyecto_id=proyecto_id)
    if fecha_desde:
        solicitudes = solicitudes.filter(fecha_solicitud__gte=fecha_desde)
    if fecha_hasta:
        solicitudes = solicitudes.filter(fecha_solicitud__lte=fecha_hasta)
    if con_ticket == 'si':
        solicitudes = solicitudes.filter(ticket__isnull=False)
    elif con_ticket == 'no':
        solicitudes = solicitudes.filter(ticket__isnull=True)

    if sin_ticket == 'si':
        solicitudes = solicitudes.filter(ticket__isnull=True)
    
    # Paginación
    por_pagina = request.GET.get('por_pagina', 10)
    try:
        por_pagina = int(por_pagina)
    except ValueError:
        por_pagina = 20
    
    paginator = Paginator(solicitudes, por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'solicitudes': page_obj,
        'page_obj': page_obj,
        'clientes': Cliente.objects.filter(activo=True),
        'proyectos': Proyecto.objects.filter(activo=True),
        'total_solicitudes': SolicitudPruebas.objects.count(),
        'solicitudes_con_ticket': SolicitudPruebas.objects.filter(ticket__isnull=False).count(),
        'solicitudes_sin_ticket': SolicitudPruebas.objects.filter(ticket__isnull=True).count(),
        # Filtros actuales
        'cliente_selected': int(cliente_id) if cliente_id else 0,
        'proyecto_selected': int(proyecto_id) if proyecto_id else 0,
        'fecha_desde': fecha_desde or '',
        'fecha_hasta': fecha_hasta or '',
        'con_ticket': con_ticket or '',
        'por_pagina': por_pagina,
        'today': today,
        'week_ago': week_ago,
    }
    return render(request, 'catalogos/solicitud_list.html', context)


def solicitud_detail(request, id):
    """Ver detalle de una solicitud de pruebas"""
    solicitud = get_object_or_404(SolicitudPruebas, id=id)
    
    context = {
        'solicitud': solicitud,
    }
    return render(request, 'catalogos/solicitud_detail.html', context)


def solicitud_generar_ticket(request, id):
    """Generar un ticket a partir de una solicitud existente"""
    solicitud = get_object_or_404(SolicitudPruebas, id=id)
    
    if request.method == 'POST':
        try:
            if solicitud.ticket:
                messages.warning(request, f'Esta solicitud ya tiene un ticket asociado: {solicitud.ticket.codigo}')
                return redirect('extractor:eticket_detail', id=solicitud.ticket.id)
            
            ticket = solicitud.generar_ticket()
            messages.success(request, f'✅ Ticket generado exitosamente: {ticket.codigo}')
            return redirect('extractor:eticket_detail', id=ticket.id)
            
        except Exception as e:
            messages.error(request, f'Error al generar ticket: {str(e)}')
            return redirect('extractor:solicitud_detail', id=solicitud.id)
    
    # GET - Mostrar confirmación
    context = {
        'solicitud': solicitud,
    }
    return render(request, 'catalogos/solicitud_generar_ticket.html', context)

def solicitud_delete(request, id):
    """Eliminar una solicitud de pruebas"""
    solicitud = get_object_or_404(SolicitudPruebas, id=id)
    
    if request.method == 'POST':
        try:
            if solicitud.ticket:
                messages.error(request, 'No se puede eliminar una solicitud que tiene un ticket asociado')
                return redirect('extractor:solicitud_detail', id=solicitud.id)
            
            solicitud.delete()
            messages.success(request, '✅ Solicitud eliminada exitosamente')
            return redirect('extractor:solicitud_list')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar solicitud: {str(e)}')
            return redirect('extractor:solicitud_detail', id=solicitud.id)
    
    # GET - Mostrar confirmación
    context = {
        'solicitud': solicitud,
    }
    return render(request, 'catalogos/solicitud_confirm_delete.html', context)

def imprimir_solicitud_excel(request, id):
    """
    Genera el archivo Excel de solicitud de pruebas usando la plantilla
    """
    import io
    import os
    from datetime import datetime
    from openpyxl import load_workbook
    from django.conf import settings
    
    solicitud = get_object_or_404(SolicitudPruebas, id=id)
    
    # Ruta a la plantilla
    plantilla_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'plantillas',
        'BID-PMC-FOR-00017_Formato_de_Solicitud_de_Pruebas.xlsx'
    )
    
    if not os.path.exists(plantilla_path):
        messages.error(request, f"No se encontró la plantilla en: {plantilla_path}")
        return redirect('extractor:solicitud_detail', id=solicitud.id)
    
    try:
        wb = load_workbook(plantilla_path)
        
        if 'Solicitud de Pruebas V4' in wb.sheetnames:
            ws = wb['Solicitud de Pruebas V4']
        else:
            ws = wb.active
        
        def set_cell_value(sheet, coordinate, value):
            """Escribe un valor en una celda, manejando correctamente celdas fusionadas"""
            try:
                for merged_range in sheet.merged_cells.ranges:
                    if coordinate in merged_range:
                        top_left = merged_range.start_cell.coordinate
                        sheet[top_left] = value
                        return
                sheet[coordinate] = value
            except Exception as e:
                print(f"⚠️ Error en {coordinate}: {e}")
        
        # ===== LLENAR DATOS DE LA SOLICITUD =====
        
        # Cliente (C5)
        if solicitud.cliente:
            set_cell_value(ws, 'C5', solicitud.cliente.nombre)
        
        # Proyecto (H5)
        if solicitud.proyecto:
            set_cell_value(ws, 'H5', solicitud.proyecto.nombre)
        
        # Fecha Solicitud (M5) - CORREGIDO
        if solicitud.fecha_solicitud:
            if hasattr(solicitud.fecha_solicitud, 'strftime'):
                fecha_str = solicitud.fecha_solicitud.strftime('%d/%m/%Y')
            else:
                try:
                    fecha_obj = datetime.strptime(str(solicitud.fecha_solicitud), '%Y-%m-%d')
                    fecha_str = fecha_obj.strftime('%d/%m/%Y')
                except:
                    fecha_str = str(solicitud.fecha_solicitud)
            set_cell_value(ws, 'M5', fecha_str)
        
        # Hora Solicitud (M6) - CORREGIDO
        if solicitud.hora_solicitud:
            if hasattr(solicitud.hora_solicitud, 'strftime'):
                hora_str = solicitud.hora_solicitud.strftime('%H:%M') + ' hrs'
            else:
                try:
                    hora_obj = datetime.strptime(str(solicitud.hora_solicitud), '%H:%M:%S')
                    hora_str = hora_obj.strftime('%H:%M') + ' hrs'
                except:
                    try:
                        hora_obj = datetime.strptime(str(solicitud.hora_solicitud), '%H:%M')
                        hora_str = hora_obj.strftime('%H:%M') + ' hrs'
                    except:
                        hora_str = str(solicitud.hora_solicitud) + ' hrs'
            set_cell_value(ws, 'M6', hora_str)
        
        # Tipo de Pruebas (D8)
        if solicitud.tipo_prueba:
            set_cell_value(ws, 'D8', solicitud.tipo_prueba.nombre)
        
        # Área Solicitante (K8)
        set_cell_value(ws, 'K8', solicitud.area_solicitante or '')
        
        # Responsable Solicitud (D12)
        set_cell_value(ws, 'D12', solicitud.responsable_solicitud or '')
        
        # Líder de Proyecto (J12)
        set_cell_value(ws, 'J12', solicitud.lider_proyecto or '')
        
        # Tipo de Aplicación (D17)
        set_cell_value(ws, 'D17', solicitud.tipo_aplicacion or '')
        
        # Número de Versión (M17)
        set_cell_value(ws, 'M17', solicitud.numero_version or '')
        
        # Funcionalidad de la liberación (D20)
        if solicitud.funcionalidad_liberacion:
            set_cell_value(ws, 'D20', solicitud.funcionalidad_liberacion)
        
        # Detalle de los cambios (D22)
        if solicitud.detalle_cambios:
            set_cell_value(ws, 'D22', solicitud.detalle_cambios)
        
        # Justificación del cambio (D24)
        if solicitud.justificacion_cambio:
            set_cell_value(ws, 'D24', solicitud.justificacion_cambio)
        
        # Puntos a considerar (D26)
        if solicitud.puntos_considerar:
            set_cell_value(ws, 'D26', solicitud.puntos_considerar)
        
        # Pendientes (D28)
        if solicitud.pendientes:
            set_cell_value(ws, 'D28', solicitud.pendientes)
        
        # Insumos requeridos (D30)
        if solicitud.insumos:
            set_cell_value(ws, 'D30', solicitud.insumos)
        
        # Nombre de servicio (D37)
        set_cell_value(ws, 'D37', 'Servicio de Pruebas')
        
        # Soporte back (J37)
        set_cell_value(ws, 'J37', solicitud.responsable_solicitud or '')
        
        # Detalles del servicio (D39)
        detalles = f"Cliente: {solicitud.cliente.nombre if solicitud.cliente else ''} - Proyecto: {solicitud.proyecto.nombre if solicitud.proyecto else ''}"
        set_cell_value(ws, 'D39', detalles)
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Crear respuesta
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo
        if solicitud.nombre_archivo:
            filename = solicitud.nombre_archivo
        else:
            if solicitud.ticket:
                filename = f"{solicitud.ticket.codigo} Solicitud de Pruebas.xlsx"
            else:
                filename = f"Solicitud_{solicitud.id}_{solicitud.fecha_solicitud}.xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(buffer.getvalue())
        
        return response
        
    except Exception as e:
        print(f"❌ Error al generar solicitud: {str(e)}")
        import traceback
        traceback.print_exc()
        
        messages.error(request, f"Error al generar el archivo: {str(e)}")
        return redirect('extractor:solicitud_detail', id=solicitud.id)
    
def check_rate_limit_by_ip(request, limite=5, tiempo_ventana=3600):
    """
    Limita las solicitudes por IP
    - limite: 5 solicitudes máximo
    - tiempo_ventana: 1 hora (3600 segundos)
    """
    # Obtener IP del cliente
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    
    # Clave única para esta IP
    cache_key = f'rate_limit_ip_{ip}'
    
    # Obtener solicitudes actuales
    solicitudes = cache.get(cache_key, [])
    
    # Limpiar solicitudes fuera de la ventana
    tiempo_actual = timezone.now().timestamp()
    solicitudes = [s for s in solicitudes if tiempo_actual - s < tiempo_ventana]
    
    # Verificar límite
    if len(solicitudes) >= limite:
        tiempo_restante = int(tiempo_ventana - (tiempo_actual - solicitudes[0]))
        minutos = tiempo_restante // 60
        return False, f"Has alcanzado el límite de 5 solicitudes por hora. Espera {minutos} minutos."
    
    # Agregar nueva solicitud
    solicitudes.append(tiempo_actual)
    cache.set(cache_key, solicitudes, timeout=tiempo_ventana)
    
    return True, ""

def registro_view(request):
    """Vista para registro de nuevos usuarios"""
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Iniciar sesión automáticamente
            login(request, user)
            messages.success(request, f'¡Bienvenido {user.username}! Tu cuenta ha sido creada exitosamente.')
            return redirect('extractor:solicitud_list') 
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = RegistroUsuarioForm()
    
    return render(request, 'extractor/registro.html', {'form': form})

# ===== VISTAS PARA GESTIÓN DE USUARIOS =====

@login_required
def usuarios_list(request):
    """
    Lista todos los usuarios registrados en el sistema
    Solo accesible para superusuarios o usuarios con permisos especiales
    """
    # Verificar permisos (solo admin puede ver lista de usuarios)
    if not request.user.is_superuser and not request.user.has_perm('auth.view_user'):
        messages.error(request, 'No tienes permiso para ver la lista de usuarios')
        return redirect('extractor:solicitud_list')
    
    usuarios = Usuario.objects.all().select_related('cliente_asociado')
    
    # Filtros
    rol = request.GET.get('rol')
    cliente_id = request.GET.get('cliente')
    search = request.GET.get('q')
    is_active = request.GET.get('activo')
    
    if rol:
        if rol == 'admin':
            usuarios = usuarios.filter(is_superuser=True)
        elif rol == 'staff':
            usuarios = usuarios.filter(is_staff=True, is_superuser=False)
        elif rol == 'user':
            usuarios = usuarios.filter(is_staff=False, is_superuser=False)
    
    if cliente_id:
        usuarios = usuarios.filter(cliente_asociado_id=cliente_id)
    
    if search:
        usuarios = usuarios.filter(
            Q(username__icontains=search) |
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(puesto__icontains=search)
        )
    
    if is_active == 'si':
        usuarios = usuarios.filter(is_active=True)
    elif is_active == 'no':
        usuarios = usuarios.filter(is_active=False)
    
    # Ordenamiento
    orden = request.GET.get('orden', '-date_joined')
    usuarios = usuarios.order_by(orden)
    
    # Paginación
    por_pagina = request.GET.get('por_pagina', 20)
    try:
        por_pagina = int(por_pagina)
    except ValueError:
        por_pagina = 20
    
    paginator = Paginator(usuarios, por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    context = {
        'usuarios': page_obj,
        'page_obj': page_obj,
        'total_usuarios': Usuario.objects.count(),
        'usuarios_activos': Usuario.objects.filter(is_active=True).count(),
        'usuarios_inactivos': Usuario.objects.filter(is_active=False).count(),
        'admins': Usuario.objects.filter(is_superuser=True).count(),
        'staff': Usuario.objects.filter(is_staff=True, is_superuser=False).count(),
        'clientes': Cliente.objects.filter(activo=True),
        'filtro_rol': rol,
        'filtro_cliente': cliente_id,
        'filtro_activo': is_active,
        'busqueda': search or '',
        'orden_actual': orden,
        'por_pagina': por_pagina,
    }
    return render(request, 'catalogos/usuarios_list.html', context)


@login_required
def usuario_detail(request, id):
    """Ver detalle de un usuario específico"""
    if not request.user.is_superuser and request.user.id != id:
        messages.error(request, 'No tienes permiso para ver este perfil')
        return redirect('extractor:usuarios_list')
    
    usuario = get_object_or_404(Usuario, id=id)
    
    # Obtener tickets creados por el usuario
    tickets_creados = Ticket.objects.filter(creado_por=usuario).order_by('-fecha_creacion')[:10]
    
    # Obtener tickets asignados al usuario
    tickets_asignados = Ticket.objects.filter(asignado_a=usuario).order_by('-fecha_creacion')[:10]
    
    context = {
        'usuario': usuario,
        'tickets_creados': tickets_creados,
        'tickets_asignados': tickets_asignados,
        'total_tickets_creados': Ticket.objects.filter(creado_por=usuario).count(),
        'total_tickets_asignados': Ticket.objects.filter(asignado_a=usuario).count(),
    }
    return render(request, 'catalogos/usuario_detail.html', context)


@login_required
def usuario_edit(request, id):
    """Editar un usuario existente"""
    if not request.user.is_superuser and request.user.id != id:
        messages.error(request, 'No tienes permiso para editar este usuario')
        return redirect('extractor:usuarios_list')
    
    usuario = get_object_or_404(Usuario, id=id)
    
    if request.method == 'POST':
        try:
            # Datos básicos
            usuario.first_name = request.POST.get('first_name', '')
            usuario.last_name = request.POST.get('last_name', '')
            usuario.email = request.POST.get('email', '')
            usuario.telefono = request.POST.get('telefono', '')
            usuario.puesto = request.POST.get('puesto', '')
            
            cliente_id = request.POST.get('cliente_asociado')
            if cliente_id:
                usuario.cliente_asociado_id = cliente_id
            else:
                usuario.cliente_asociado = None
            
            # Permisos (solo superusuario puede cambiar estos)
            if request.user.is_superuser:
                usuario.is_active = request.POST.get('is_active', 'off') == 'on'
                usuario.is_staff = request.POST.get('is_staff', 'off') == 'on'
                usuario.is_superuser = request.POST.get('is_superuser', 'off') == 'on'
                usuario.puede_generar_tickets = request.POST.get('puede_generar_tickets', 'off') == 'on'
                usuario.puede_ver_todos_tickets = request.POST.get('puede_ver_todos_tickets', 'off') == 'on'
            
            # Cambiar contraseña si se proporcionó
            nueva_password = request.POST.get('new_password')
            if nueva_password:
                if len(nueva_password) >= 8:
                    usuario.set_password(nueva_password)
                    messages.success(request, 'Contraseña actualizada exitosamente')
                else:
                    messages.error(request, 'La contraseña debe tener al menos 8 caracteres')
                    return redirect('extractor:usuario_edit', id=usuario.id)
            
            usuario.save()
            messages.success(request, f'Usuario "{usuario.username}" actualizado exitosamente')
            
            if request.user.is_superuser:
                return redirect('extractor:usuarios_list')
            else:
                return redirect('extractor:usuario_detail', id=usuario.id)
                
        except Exception as e:
            messages.error(request, f'Error al actualizar usuario: {str(e)}')
            return redirect('extractor:usuario_edit', id=usuario.id)
    
    context = {
        'usuario': usuario,
        'clientes': Cliente.objects.filter(activo=True),
        'es_superusuario': request.user.is_superuser,
    }
    return render(request, 'catalogos/usuario_form.html', context)


@login_required
def usuario_create(request):
    """Crear un nuevo usuario manualmente (solo superusuarios)"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permiso para crear usuarios')
        return redirect('extractor:usuarios_list')
    
    if request.method == 'POST':
        try:
            username = request.POST.get('username', '').strip()
            email = request.POST.get('email', '').strip()
            password = request.POST.get('password', '')
            password_confirm = request.POST.get('password_confirm', '')
            
            # Validaciones
            if not username or not email:
                messages.error(request, 'Usuario y email son obligatorios')
                return redirect('extractor:usuario_create')
            
            if Usuario.objects.filter(username=username).exists():
                messages.error(request, f'El usuario "{username}" ya existe')
                return redirect('extractor:usuario_create')
            
            if Usuario.objects.filter(email=email).exists():
                messages.error(request, f'El email "{email}" ya está registrado')
                return redirect('extractor:usuario_create')
            
            if password != password_confirm:
                messages.error(request, 'Las contraseñas no coinciden')
                return redirect('extractor:usuario_create')
            
            if len(password) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres')
                return redirect('extractor:usuario_create')
            
            # Crear usuario
            usuario = Usuario.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=request.POST.get('first_name', ''),
                last_name=request.POST.get('last_name', ''),
                telefono=request.POST.get('telefono', ''),
                puesto=request.POST.get('puesto', ''),
            )
            
            # Asignar cliente
            cliente_id = request.POST.get('cliente_asociado')
            if cliente_id:
                usuario.cliente_asociado_id = cliente_id
            
            # Permisos
            usuario.is_active = request.POST.get('is_active', 'on') == 'on'
            usuario.is_staff = request.POST.get('is_staff', 'off') == 'on'
            usuario.is_superuser = request.POST.get('is_superuser', 'off') == 'on'
            usuario.puede_generar_tickets = request.POST.get('puede_generar_tickets', 'on') == 'on'
            usuario.puede_ver_todos_tickets = request.POST.get('puede_ver_todos_tickets', 'off') == 'on'
            usuario.save()
            
            messages.success(request, f'✅ Usuario "{usuario.username}" creado exitosamente')
            return redirect('extractor:usuarios_list')
            
        except Exception as e:
            messages.error(request, f'Error al crear usuario: {str(e)}')
            return redirect('extractor:usuario_create')
    
    context = {
        'clientes': Cliente.objects.filter(activo=True),
    }
    return render(request, 'catalogos/usuario_create_form.html')


@login_required
def usuario_delete(request, id):
    """Eliminar (desactivar) un usuario"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permiso para eliminar usuarios')
        return redirect('extractor:usuarios_list')
    
    usuario = get_object_or_404(Usuario, id=id)
    
    # No permitir eliminar el propio usuario
    if usuario.id == request.user.id:
        messages.error(request, 'No puedes eliminarte a ti mismo')
        return redirect('extractor:usuarios_list')
    
    if request.method == 'POST':
        try:
            username = usuario.username
            # En lugar de eliminar, desactivar (soft delete)
            usuario.is_active = False
            usuario.save()
            messages.success(request, f'Usuario "{username}" desactivado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al desactivar usuario: {str(e)}')
    
    return redirect('extractor:usuarios_list')


@login_required
def usuario_activar(request, id):
    """Reactivar un usuario desactivado"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permiso para activar usuarios')
        return redirect('extractor:usuarios_list')
    
    usuario = get_object_or_404(Usuario, id=id)
    
    if request.method == 'POST':
        try:
            usuario.is_active = True
            usuario.save()
            messages.success(request, f'Usuario "{usuario.username}" activado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al activar usuario: {str(e)}')
    
    return redirect('extractor:usuarios_list')


@login_required
def usuario_cambiar_rol(request, id):
    """Cambiar rol de usuario (API AJAX)"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permiso denegado'})
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            usuario = get_object_or_404(Usuario, id=id)
            nuevo_rol = data.get('rol')
            
            if nuevo_rol == 'admin':
                usuario.is_superuser = True
                usuario.is_staff = True
            elif nuevo_rol == 'staff':
                usuario.is_superuser = False
                usuario.is_staff = True
            elif nuevo_rol == 'user':
                usuario.is_superuser = False
                usuario.is_staff = False
            else:
                return JsonResponse({'success': False, 'error': 'Rol no válido'})
            
            usuario.save()
            
            return JsonResponse({
                'success': True,
                'rol_display': usuario.get_rol_display() if hasattr(usuario, 'get_rol_display') else nuevo_rol
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def export_usuarios_csv(request):
    """Exportar usuarios a CSV"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permiso para exportar usuarios')
        return redirect('extractor:usuarios_list')
    
    try:
        usuarios = Usuario.objects.all().select_related('cliente_asociado')
        
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))
        
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"usuarios_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Usuario', 'Email', 'Nombre', 'Apellido', 'Teléfono', 
            'Puesto', 'Cliente Asociado', 'Activo', 'Staff', 'Superusuario',
            'Puede Generar Tickets', 'Puede Ver Todos Tickets', 'Fecha Registro'
        ])
        
        for usuario in usuarios:
            writer.writerow([
                usuario.id,
                usuario.username,
                usuario.email,
                usuario.first_name,
                usuario.last_name,
                usuario.telefono or '',
                usuario.puesto or '',
                usuario.cliente_asociado.nombre if usuario.cliente_asociado else '',
                'Sí' if usuario.is_active else 'No',
                'Sí' if usuario.is_staff else 'No',
                'Sí' if usuario.is_superuser else 'No',
                'Sí' if usuario.puede_generar_tickets else 'No',
                'Sí' if usuario.puede_ver_todos_tickets else 'No',
                usuario.date_joined.strftime('%d/%m/%Y %H:%M') if usuario.date_joined else ''
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exportando usuarios: {str(e)}", exc_info=True)
        messages.error(request, "Error al exportar usuarios")
        return redirect('extractor:usuarios_list')

@login_required
def export_clientes_csv(request):
    """Exporta clientes a CSV"""
    try:
        clientes = Cliente.objects.all()
        
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))
        
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"clientes_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Nombre', 'Nomenclatura', 'Activo', 'Fecha Creación'])
        
        for cliente in clientes:
            writer.writerow([
                cliente.id,
                cliente.nombre,
                cliente.nomenclatura,
                'Sí' if cliente.activo else 'No',
                cliente.fecha_creacion.strftime('%d/%m/%Y %H:%M') if cliente.fecha_creacion else ''
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exportando clientes: {str(e)}")
        messages.error(request, "Error al exportar clientes")
        return redirect('extractor:clientes_list')


@login_required
def export_proyectos_csv(request):
    """Exporta proyectos a CSV"""
    try:
        proyectos = Proyecto.objects.all().select_related('cliente')
        
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))
        
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"proyectos_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Cliente', 'Nombre', 'Código', 'Descripción', 'Activo', 'Fecha Inicio', 'Fecha Fin'])
        
        for proyecto in proyectos:
            writer.writerow([
                proyecto.id,
                proyecto.cliente.nombre if proyecto.cliente else '',
                proyecto.nombre,
                proyecto.codigo,
                proyecto.descripcion or '',
                'Sí' if proyecto.activo else 'No',
                proyecto.fecha_inicio.strftime('%d/%m/%Y') if proyecto.fecha_inicio else '',
                proyecto.fecha_fin.strftime('%d/%m/%Y') if proyecto.fecha_fin else ''
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exportando proyectos: {str(e)}")
        messages.error(request, "Error al exportar proyectos")
        return redirect('extractor:proyectos_list')


@login_required
def export_tipos_servicio_csv(request):
    """Exporta tipos de servicio a CSV"""
    try:
        tipos = TipoServicio.objects.all()
        
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))
        
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"tipos_servicio_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Nombre', 'Nomenclatura', 'Activo', 'Fecha Creación'])
        
        for tipo in tipos:
            writer.writerow([
                tipo.id,
                tipo.nombre,
                tipo.nomenclatura,
                'Sí' if tipo.activo else 'No',
                tipo.fecha_creacion.strftime('%d/%m/%Y %H:%M') if tipo.fecha_creacion else ''
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exportando tipos de servicio: {str(e)}")
        messages.error(request, "Error al exportar tipos de servicio")
        return redirect('extractor:tipo_servicio_list')


@login_required
def export_all_tables_backup(request):
    """Exporta todas las tablas como CSV en un archivo ZIP"""
    import zipfile
    import io
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    
    try:
        # Crear archivo ZIP en memoria
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            models_to_export = {
                'clientes': Cliente,
                'proyectos': Proyecto,
                'tipos_servicio': TipoServicio,
                'tickets': Ticket,
                'datos_excel': ExcelData,
                'solicitudes_pruebas': SolicitudPruebas,
                'usuarios': Usuario,
            }
            
            for filename, model in models_to_export.items():
                # Crear CSV en memoria
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                
                queryset = model.objects.all()
                
                # Escribir encabezados
                headers = [field.name for field in model._meta.fields]
                writer.writerow(headers)
                
                # Escribir datos
                for obj in queryset:
                    row = []
                    for field in headers:
                        value = getattr(obj, field)
                        if value is None:
                            row.append('')
                        elif hasattr(value, 'strftime'):
                            row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                        elif hasattr(value, 'pk'):
                            row.append(value.pk)
                        else:
                            row.append(str(value))
                    writer.writerow(row)
                
                # Convertir StringIO a bytes para el ZIP
                csv_content = csv_buffer.getvalue().encode('utf-8-sig')
                zip_file.writestr(f"{filename}.csv", csv_content)
        
        # Preparar respuesta
        zip_buffer.seek(0)
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="backup_completo_{timestamp}.zip"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error en backup: {str(e)}", exc_info=True)
        messages.error(request, f"Error al crear backup: {str(e)}")
        return redirect('extractor:solicitud_list')


@login_required
def subir_dictamen(request, id):
    """Subir archivo PDF del dictamen a Cloudinary cuando el ticket está COMPLETADO"""
    ticket = get_object_or_404(Ticket, id=id)
    
    # Verificar que el ticket esté COMPLETADO
    if ticket.estado != 'COMPLETADO':
        messages.error(request, 'Solo se pueden subir archivos cuando el ticket está COMPLETADO')
        return redirect('extractor:ticket_detail', id=ticket.id)
    
    if request.method == 'POST' and request.FILES.get('dictamen_pdf'):
        archivo = request.FILES['dictamen_pdf']
        
        # Validar que sea PDF
        if not archivo.name.endswith('.pdf'):
            messages.error(request, 'Solo se permiten archivos PDF')
            return redirect('extractor:ticket_detail', id=ticket.id)
        
        # Validar tamaño máximo (10MB)
        if archivo.size > 10 * 1024 * 1024:
            messages.error(request, 'El archivo no puede superar los 10MB')
            return redirect('extractor:ticket_detail', id=ticket.id)
        
        try:
            import cloudinary.uploader
            from django.utils import timezone
            
            # Subir a Cloudinary
            resultado = cloudinary.uploader.upload(
                archivo,
                folder=f"tickets/{ticket.id}/dictamenes",
                resource_type="auto",
                allowed_formats=["pdf"],
                public_id=f"dictamen_{ticket.codigo.replace('-', '_')}"
            )
            
            # Guardar la URL en el ticket
            ticket.dictamen_pdf = resultado['secure_url']
            ticket.dictamen_url = resultado['secure_url']
            ticket.fecha_subida_dictamen = timezone.now()
            ticket.subido_por = request.user
            ticket.save()
            
            # Registrar comentario automático
            usuario = request.user.get_full_name() or request.user.username
            ahora_local = timezone.localtime(timezone.now())
            fecha_hora = ahora_local.strftime('%d/%m/%Y %H:%M')
            comentario = f"[{fecha_hora}] {usuario} subió el dictamen: {archivo.name}"
            
            if ticket.comentarios_seguimiento:
                ticket.comentarios_seguimiento += f"\n{comentario}"
            else:
                ticket.comentarios_seguimiento = comentario
            ticket.save()
            
            messages.success(request, f'✅ Dictamen subido exitosamente a Cloudinary: {archivo.name}')
            
        except Exception as e:
            messages.error(request, f'Error al subir a Cloudinary: {str(e)}')
    
    return redirect('extractor:ticket_detail', id=ticket.id)


@login_required
def subir_evidencia(request, id):
    """Subir archivo PDF de evidencia a Cloudinary cuando el ticket está COMPLETADO"""
    ticket = get_object_or_404(Ticket, id=id)
    
    if ticket.estado != 'COMPLETADO':
        messages.error(request, 'Solo se pueden subir archivos cuando el ticket está COMPLETADO')
        return redirect('extractor:ticket_detail', id=ticket.id)
    
    if request.method == 'POST' and request.FILES.get('evidencia_pdf'):
        archivo = request.FILES['evidencia_pdf']
        
        if not archivo.name.endswith('.pdf'):
            messages.error(request, 'Solo se permiten archivos PDF')
            return redirect('extractor:ticket_detail', id=ticket.id)
        
        if archivo.size > 10 * 1024 * 1024:
            messages.error(request, 'El archivo no puede superar los 10MB')
            return redirect('extractor:ticket_detail', id=ticket.id)
        
        try:
            import cloudinary.uploader
            from django.utils import timezone
            
            resultado = cloudinary.uploader.upload(
                archivo,
                folder=f"tickets/{ticket.id}/evidencias",
                resource_type="auto",
                allowed_formats=["pdf"],
                public_id=f"evidencia_{ticket.codigo.replace('-', '_')}"
            )
            
            ticket.evidencia_pdf = resultado['secure_url']
            ticket.evidencia_url = resultado['secure_url']
            ticket.fecha_subida_evidencia = timezone.now()
            ticket.subido_por = request.user
            ticket.save()
            
            usuario = request.user.get_full_name() or request.user.username
            ahora_local = timezone.localtime(timezone.now())
            fecha_hora = ahora_local.strftime('%d/%m/%Y %H:%M')
            comentario = f"[{fecha_hora}] {usuario} subió evidencia de pruebas: {archivo.name}"
            
            if ticket.comentarios_seguimiento:
                ticket.comentarios_seguimiento += f"\n{comentario}"
            else:
                ticket.comentarios_seguimiento = comentario
            ticket.save()
            
            messages.success(request, f'✅ Evidencia subida exitosamente a Cloudinary: {archivo.name}')
            
        except Exception as e:
            messages.error(request, f'Error al subir a Cloudinary: {str(e)}')
    
    return redirect('extractor:ticket_detail', id=ticket.id)


@login_required
def eliminar_archivo_cloudinary(request, id, tipo):
    """Eliminar archivo de Cloudinary"""
    ticket = get_object_or_404(Ticket, id=id)
    
    if ticket.estado != 'COMPLETADO':
        messages.error(request, 'Solo se pueden eliminar archivos de tickets COMPLETADOS')
        return redirect('extractor:ticket_detail', id=ticket.id)
    
    try:
        import cloudinary.uploader
        from django.utils import timezone
        
        usuario = request.user.get_full_name() or request.user.username
        ahora_local = timezone.localtime(timezone.now())
        fecha_hora = ahora_local.strftime('%d/%m/%Y %H:%M')
        
        if tipo == 'dictamen' and ticket.dictamen_pdf:
            public_id = f"tickets/{ticket.id}/dictamenes/dictamen_{ticket.codigo.replace('-', '_')}"
            cloudinary.uploader.destroy(public_id, resource_type="image")
            
            ticket.dictamen_pdf = None
            ticket.dictamen_url = None
            ticket.fecha_subida_dictamen = None
            comentario = f"[{fecha_hora}] {usuario} eliminó el dictamen"
            messages.success(request, '✅ Dictamen eliminado de Cloudinary')
            
        elif tipo == 'evidencia' and ticket.evidencia_pdf:
            public_id = f"tickets/{ticket.id}/evidencias/evidencia_{ticket.codigo.replace('-', '_')}"
            cloudinary.uploader.destroy(public_id, resource_type="image")
            
            ticket.evidencia_pdf = None
            ticket.evidencia_url = None
            ticket.fecha_subida_evidencia = None
            comentario = f"[{fecha_hora}] {usuario} eliminó la evidencia"
            messages.success(request, '✅ Evidencia eliminada de Cloudinary')
        else:
            messages.error(request, 'Archivo no encontrado')
            return redirect('extractor:ticket_detail', id=ticket.id)
        
        ticket.save()
        
        if ticket.comentarios_seguimiento:
            ticket.comentarios_seguimiento += f"\n{comentario}"
        else:
            ticket.comentarios_seguimiento = comentario
        ticket.save()
        
    except Exception as e:
        messages.error(request, f'Error al eliminar de Cloudinary: {str(e)}')
    
    return redirect('extractor:ticket_detail', id=ticket.id)