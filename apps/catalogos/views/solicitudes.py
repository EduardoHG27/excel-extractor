"""
Vistas para gestión de Solicitudes de Pruebas
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.core.cache import cache
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
import io
import os
import csv
from datetime import datetime
from openpyxl import load_workbook
from django.conf import settings

from extractor.models import Cliente, Proyecto, TipoServicio, SolicitudPruebas, Ticket


def check_rate_limit_by_ip(request, limite=5, tiempo_ventana=3600):
    """Limita las solicitudes por IP"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    
    cache_key = f'rate_limit_ip_{ip}'
    solicitudes = cache.get(cache_key, [])
    
    tiempo_actual = timezone.now().timestamp()
    solicitudes = [s for s in solicitudes if tiempo_actual - s < tiempo_ventana]
    
    if len(solicitudes) >= limite:
        tiempo_restante = int(tiempo_ventana - (tiempo_actual - solicitudes[0]))
        minutos = tiempo_restante // 60
        return False, f"Has alcanzado el límite de {limite} solicitudes por hora. Espera {minutos} minutos."
    
    solicitudes.append(tiempo_actual)
    cache.set(cache_key, solicitudes, timeout=tiempo_ventana)
    
    return True, ""


@login_required
def solicitud_list(request):
    """Listado de solicitudes de pruebas"""
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)

    solicitudes = SolicitudPruebas.objects.all().select_related('cliente', 'proyecto', 'tipo_prueba', 'ticket')
    
    # Filtros
    cliente_id = request.GET.get('cliente')
    proyecto_id = request.GET.get('proyecto')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    con_ticket = request.GET.get('con_ticket')
    
    if cliente_id:
        solicitudes = solicitudes.filter(cliente_id=cliente_id)
    if proyecto_id:
        solicitudes = solicitudes.filter(proyecto_id=proyecto_id)
    if fecha_desde:
        solicitudes = solicitudes.filter(fecha_solicitud__gte=fecha_desde)
    if fecha_hasta:
        solicitudes = solicitudes.filter(fecha_solicitud__lte=fecha_hasta)
    if con_ticket == 'si':
        solicitudes = solicitudes.filter(ticket__isnull=False)
    elif con_ticket == 'no':
        solicitudes = solicitudes.filter(ticket__isnull=True)
    
    # Paginación
    por_pagina = request.GET.get('por_pagina', 10)
    try:
        por_pagina = int(por_pagina)
    except ValueError:
        por_pagina = 20
    
    paginator = Paginator(solicitudes, por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'solicitudes': page_obj,
        'page_obj': page_obj,
        'clientes': Cliente.objects.filter(activo=True),
        'proyectos': Proyecto.objects.filter(activo=True),
        'total_solicitudes': SolicitudPruebas.objects.count(),
        'solicitudes_con_ticket': SolicitudPruebas.objects.filter(ticket__isnull=False).count(),
        'solicitudes_sin_ticket': SolicitudPruebas.objects.filter(ticket__isnull=True).count(),
        'cliente_selected': int(cliente_id) if cliente_id else 0,
        'proyecto_selected': int(proyecto_id) if proyecto_id else 0,
        'fecha_desde': fecha_desde or '',
        'fecha_hasta': fecha_hasta or '',
        'con_ticket': con_ticket or '',
        'por_pagina': por_pagina,
        'today': today,
        'week_ago': week_ago,
        'debug': settings.DEBUG,
    }
    return render(request, 'catalogos/solicitud_list.html', context)


@login_required
def solicitud_detail(request, id):
    """Ver detalle de una solicitud de pruebas"""
    solicitud = get_object_or_404(SolicitudPruebas, id=id)
    
    context = {
        'solicitud': solicitud,
    }
    return render(request, 'catalogos/solicitud_detail.html', context)


def crear_solicitud(request):
    """
    Vista para crear solicitud de pruebas manualmente
    Con protecciones anti-bots (Cooldown + Honeypot + Rate Limiting)
    """
    from django.conf import settings
    from django.utils import timezone
    from datetime import datetime
    
    # Cooldown de 5 minutos
    ultima_solicitud = request.session.get('ultima_solicitud_timestamp')
    cooldown_segundos = getattr(settings, 'SOLICITUD_COOLDOWN_SEGUNDOS', 300)
    
    tiempo_restante = 0
    if ultima_solicitud:
        tiempo_actual = timezone.now().timestamp()
        tiempo_transcurrido = tiempo_actual - ultima_solicitud
        
        if tiempo_transcurrido < cooldown_segundos:
            tiempo_restante = cooldown_segundos - tiempo_transcurrido
            
            if request.method == 'GET':
                minutos = int(tiempo_restante // 60)
                segundos = int(tiempo_restante % 60)
                messages.info(
                    request,
                    f'⏳ Puedes crear una nueva solicitud en {minutos} minutos y {segundos} segundos.'
                )
            
            if request.method == 'POST':
                minutos = int(tiempo_restante // 60)
                segundos = int(tiempo_restante % 60)
                messages.warning(
                    request, 
                    f'⏳ Debes esperar {minutos} minutos y {segundos} segundos antes de crear otra solicitud.'
                )
                return redirect('extractor:solicitud_list')
    
    # Honeypot (solo en POST)
    if request.method == 'POST':
        if request.POST.get('web_contacto', ''):
            messages.error(request, 'Actividad sospechosa detectada.')
            return redirect('extractor:crear_solicitud')
        
        if request.POST.get('confirmar_email', ''):
            messages.error(request, 'Actividad sospechosa detectada.')
            return redirect('extractor:crear_solicitud')
    
    # Rate limiting por IP
    if request.method == 'POST':
        permitido, mensaje = check_rate_limit_by_ip(request, limite=5, tiempo_ventana=3600)
        if not permitido:
            messages.error(request, mensaje)
            return redirect('extractor:solicitud_list')
    
    # Procesar formulario
    if request.method == 'POST':
        try:
            cliente_id = request.POST.get('cliente')
            proyecto_id = request.POST.get('proyecto')
            tipo_servicio_code = request.POST.get('tipo_servicio_code')
            tipo_prueba_id = request.POST.get('tipo_prueba')
            
            if not cliente_id or not proyecto_id or not tipo_servicio_code or not tipo_prueba_id:
                messages.error(request, 'Los campos obligatorios deben estar llenos')
                return redirect('extractor:crear_solicitud')
            
            cliente = Cliente.objects.get(id=cliente_id, activo=True)
            proyecto = Proyecto.objects.get(id=proyecto_id, activo=True)
            tipo_prueba = TipoServicio.objects.get(id=tipo_prueba_id, activo=True)
            
            if proyecto.cliente_id != cliente.id:
                messages.error(request, 'El proyecto no pertenece al cliente seleccionado')
                return redirect('extractor:crear_solicitud')
            
            # ========== PROCESAMIENTO CORRECTO DE FECHA Y HORA ==========
            # Obtener la hora local actual del servidor (México)
            ahora_local = timezone.localtime(timezone.now())
            
            # Procesar fecha
            fecha_str = request.POST.get('fecha_solicitud')
            if fecha_str:
                fecha_solicitud = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            else:
                fecha_solicitud = ahora_local.date()
            
            # Procesar hora - AHORA CON ZONA HORARIA CORRECTA
            hora_str = request.POST.get('hora_solicitud')
            hora_actual_local = ahora_local.time().replace(microsecond=0)
            
            if hora_str:
                try:
                    # Intentar parsear la hora del formulario
                    hora_parseada = datetime.strptime(hora_str, '%H:%M').time()
                    
                    # Verificar si la hora es razonable (no difiere más de 5 minutos de la actual)
                    hoy = datetime.today()
                    hora_parseada_dt = datetime.combine(hoy, hora_parseada)
                    hora_actual_dt = datetime.combine(hoy, hora_actual_local)
                    diferencia_segundos = abs((hora_parseada_dt - hora_actual_dt).total_seconds())
                    
                    if diferencia_segundos > 300:  # Más de 5 minutos de diferencia
                        # Usar hora actual local en lugar de la del formulario
                        hora_solicitud = hora_actual_local
                        print(f"⚠️ Hora del formulario ({hora_parseada}) reemplazada por hora local ({hora_actual_local})")
                    else:
                        hora_solicitud = hora_parseada
                        
                except Exception as e:
                    # Si falla el parseo, usar hora actual local
                    print(f"⚠️ Error parseando hora {hora_str}: {e}")
                    hora_solicitud = hora_actual_local
            else:
                # Si no hay hora en el POST, usar la hora actual local
                hora_solicitud = hora_actual_local
            
            # Debug para verificar (opcional, puedes comentar en producción)
            print(f"✅ [DEBUG] Fecha guardada: {fecha_solicitud}")
            print(f"✅ [DEBUG] Hora guardada: {hora_solicitud}")
            print(f"✅ [DEBUG] Hora local actual: {hora_actual_local}")
            print(f"✅ [DEBUG] Hora UTC actual: {timezone.now().time()}")
            
            creado_por_value = request.POST.get('creado_por', '').strip()
            email_contacto_value = request.POST.get('email_contacto', '').strip()

            print("=" * 50)
            print("POST data recibido:")
            print(f"creado_por: {request.POST.get('creado_por', 'NO ENVIADO')}")
            print(f"email_contacto: {request.POST.get('email_contacto', 'NO ENVIADO')}")
            print("=" * 50)
            

            if not creado_por_value:
                # Si el usuario está autenticado, usar su nombre como fallback
                if request.user.is_authenticated:
                    creado_por_value = request.user.get_full_name() or request.user.username
                else:
                    messages.error(request, 'Por favor, ingresa tu nombre como solicitante')
                    return redirect('extractor:crear_solicitud')
                
            # Validar que el campo creado_por no esté vacío
            if not creado_por_value:
                messages.error(request, 'Por favor, ingresa tu nombre como solicitante')
                return redirect('extractor:crear_solicitud')

            # Si el usuario está autenticado pero quiere usar un nombre diferente,
            # permitir que el campo editable tenga prioridad
            if request.user.is_authenticated and not creado_por_value:
                creado_por_value = request.user.get_full_name() or request.user.username
            
            # Crear la solicitud
            solicitud = SolicitudPruebas(
                cliente=cliente,
                proyecto=proyecto,
                fecha_solicitud=fecha_solicitud,
                hora_solicitud=hora_solicitud,
                tipo_servicio_code=tipo_servicio_code,
                tipo_prueba=tipo_prueba,
                area_solicitante=request.POST.get('area_solicitante', ''),
                numero_version=request.POST.get('numero_version', ''),
                responsable_solicitud=request.POST.get('responsable_solicitud', ''),
                lider_proyecto=request.POST.get('lider_proyecto', ''),
                tipo_aplicacion=request.POST.get('tipo_aplicacion', ''),
                funcionalidad_liberacion=request.POST.get('funcionalidad_liberacion', ''),
                detalle_cambios=request.POST.get('detalle_cambios', ''),
                justificacion_cambio=request.POST.get('justificacion_cambio', ''),
                puntos_considerar=request.POST.get('puntos_considerar', ''),
                pendientes=request.POST.get('pendientes', ''),
                insumos=request.POST.get('insumos', ''),
                creado_por=creado_por_value,
                email_contacto=email_contacto_value, 
            )
            
            # Generar nombre de archivo (con manejo de errores)
            try:
                solicitud.nombre_archivo = solicitud.generar_nombre_archivo()
            except Exception as e:
                print(f"⚠️ Error en generar_nombre_archivo: {e}")
                # Usar nombre por defecto
                solicitud.nombre_archivo = f"Solicitud_{ahora_local.strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            solicitud.save()
            
            # Guardar timestamp de la última solicitud usando hora local
            request.session['ultima_solicitud_timestamp'] = timezone.now().timestamp()
            request.session['ultima_solicitud_id'] = solicitud.id
            
            context_exito = {
                'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
                'tipos_servicio': TipoServicio.objects.filter(activo=True).order_by('nombre'),
                'today': ahora_local.date(),
                'now': ahora_local,
                'tiempo_restante': int(tiempo_restante),
                'solicitud_creada': solicitud,
                'mostrar_resumen': True,
            }
            
            if request.POST.get('generar_ticket_ahora') == 'on':
                ticket = solicitud.generar_ticket()
                context_exito['ticket_generado'] = ticket
            
            return render(request, 'extractor/crear_solicitud.html', context_exito)
            
        except Cliente.DoesNotExist:
            messages.error(request, 'El cliente seleccionado no existe')
        except Proyecto.DoesNotExist:
            messages.error(request, 'El proyecto seleccionado no existe')
        except TipoServicio.DoesNotExist:
            messages.error(request, 'El tipo de prueba seleccionado no existe')
        except Exception as e:
            # Mostrar el error real para depuración
            print(f"❌ Error al crear solicitud: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error al crear solicitud: {str(e)}')
        
        return redirect('extractor:crear_solicitud')
    
    # GET - Mostrar formulario
    ahora_local = timezone.localtime(timezone.now())
    context = {
        'clientes': Cliente.objects.filter(activo=True).order_by('nombre'),
        'tipos_servicio': TipoServicio.objects.filter(activo=True).order_by('nombre'),
        'today': ahora_local.date(),
        'now': ahora_local,
        'tiempo_restante': int(tiempo_restante),
    }
    return render(request, 'extractor/crear_solicitud.html', context)


@login_required
def solicitud_generar_ticket(request, id):
    """Generar un ticket a partir de una solicitud existente"""
    solicitud = get_object_or_404(SolicitudPruebas, id=id)
    
    if request.method == 'POST':
        try:
            if solicitud.ticket:
                messages.warning(request, f'Esta solicitud ya tiene un ticket asociado: {solicitud.ticket.codigo}')
                return redirect('extractor:ticket_detail', id=solicitud.ticket.id)
            
            ticket = solicitud.generar_ticket()
            messages.success(request, f'✅ Ticket generado exitosamente: {ticket.codigo}')
            return redirect('extractor:ticket_detail', id=ticket.id)
            
        except Exception as e:
            messages.error(request, f'Error al generar ticket: {str(e)}')
            return redirect('extractor:solicitud_detail', id=solicitud.id)
    
    context = {'solicitud': solicitud}
    return render(request, 'catalogos/solicitud_generar_ticket.html', context)


@login_required
def solicitud_delete(request, id):
    """Eliminar una solicitud de pruebas"""
    solicitud = get_object_or_404(SolicitudPruebas, id=id)
    
    if request.method == 'POST':
        try:
            if solicitud.ticket:
                messages.error(request, 'No se puede eliminar una solicitud que tiene un ticket asociado')
                return redirect('extractor:solicitud_detail', id=solicitud.id)
            
            solicitud.delete()
            messages.success(request, '✅ Solicitud eliminada exitosamente')
            return redirect('extractor:solicitud_list')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar solicitud: {str(e)}')
            return redirect('extractor:solicitud_detail', id=solicitud.id)
    
    context = {'solicitud': solicitud}
    return render(request, 'catalogos/solicitud_confirm_delete.html', context)


@login_required
def imprimir_solicitud_excel(request, id):
    """Genera el archivo Excel de solicitud de pruebas usando la plantilla"""
    solicitud = get_object_or_404(SolicitudPruebas, id=id)
    
    plantilla_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'plantillas',
        'BID-PMC-FOR-00017_Formato_de_Solicitud_de_Pruebas.xlsx'
    )
    
    if not os.path.exists(plantilla_path):
        messages.error(request, f"No se encontró la plantilla en: {plantilla_path}")
        return redirect('extractor:solicitud_detail', id=solicitud.id)
    
    try:
        wb = load_workbook(plantilla_path)
        
        if 'Solicitud de Pruebas V4' in wb.sheetnames:
            ws = wb['Solicitud de Pruebas V4']
        else:
            ws = wb.active
        
        def set_cell_value(sheet, coordinate, value):
            try:
                for merged_range in sheet.merged_cells.ranges:
                    if coordinate in merged_range:
                        top_left = merged_range.start_cell.coordinate
                        sheet[top_left] = value
                        return
                sheet[coordinate] = value
            except Exception as e:
                print(f"⚠️ Error en {coordinate}: {e}")
        
        # Cliente (C5)
        if solicitud.cliente:
            set_cell_value(ws, 'C5', solicitud.cliente.nombre)
        
        # Proyecto (H5)
        if solicitud.proyecto:
            set_cell_value(ws, 'H5', solicitud.proyecto.nombre)
        
        # Fecha Solicitud (M5)
        if solicitud.fecha_solicitud:
            if hasattr(solicitud.fecha_solicitud, 'strftime'):
                fecha_str = solicitud.fecha_solicitud.strftime('%d/%m/%Y')
            else:
                try:
                    fecha_obj = datetime.strptime(str(solicitud.fecha_solicitud), '%Y-%m-%d')
                    fecha_str = fecha_obj.strftime('%d/%m/%Y')
                except:
                    fecha_str = str(solicitud.fecha_solicitud)
            set_cell_value(ws, 'M5', fecha_str)
        
        # Hora Solicitud (M6)
        if solicitud.hora_solicitud:
            if hasattr(solicitud.hora_solicitud, 'strftime'):
                hora_str = solicitud.hora_solicitud.strftime('%H:%M') + ' hrs'
            else:
                try:
                    hora_obj = datetime.strptime(str(solicitud.hora_solicitud), '%H:%M:%S')
                    hora_str = hora_obj.strftime('%H:%M') + ' hrs'
                except:
                    try:
                        hora_obj = datetime.strptime(str(solicitud.hora_solicitud), '%H:%M')
                        hora_str = hora_obj.strftime('%H:%M') + ' hrs'
                    except:
                        hora_str = str(solicitud.hora_solicitud) + ' hrs'
            set_cell_value(ws, 'M6', hora_str)
        
        # Tipo de Pruebas (D8)
        if solicitud.tipo_prueba:
            set_cell_value(ws, 'D8', solicitud.tipo_prueba.nombre)
        
        # Área Solicitante (K8)
        set_cell_value(ws, 'K8', solicitud.area_solicitante or '')
        
        # Responsable Solicitud (D12)
        set_cell_value(ws, 'D12', solicitud.responsable_solicitud or '')
        
        # Líder de Proyecto (J12)
        set_cell_value(ws, 'J12', solicitud.lider_proyecto or '')
        
        # Tipo de Aplicación (D17)
        set_cell_value(ws, 'D17', solicitud.tipo_aplicacion or '')
        
        # Número de Versión (M17)
        set_cell_value(ws, 'M17', solicitud.numero_version or '')
        
        # Funcionalidad (D20)
        if solicitud.funcionalidad_liberacion:
            set_cell_value(ws, 'D20', solicitud.funcionalidad_liberacion)
        
        # Detalle de cambios (D22)
        if solicitud.detalle_cambios:
            set_cell_value(ws, 'D22', solicitud.detalle_cambios)
        
        # Justificación (D24)
        if solicitud.justificacion_cambio:
            set_cell_value(ws, 'D24', solicitud.justificacion_cambio)
        
        # Puntos a considerar (D26)
        if solicitud.puntos_considerar:
            set_cell_value(ws, 'D26', solicitud.puntos_considerar)
        
        # Pendientes (D28)
        if solicitud.pendientes:
            set_cell_value(ws, 'D28', solicitud.pendientes)
        
        # Insumos (D30)
        if solicitud.insumos:
            set_cell_value(ws, 'D30', solicitud.insumos)
        
        # Nombre de servicio (D37)
        set_cell_value(ws, 'D37', 'Servicio de Pruebas')
        
        # Soporte back (J37)
        set_cell_value(ws, 'J37', solicitud.responsable_solicitud or '')
        
        # Detalles del servicio (D39)
        detalles = f"Cliente: {solicitud.cliente.nombre if solicitud.cliente else ''} - Proyecto: {solicitud.proyecto.nombre if solicitud.proyecto else ''}"
        set_cell_value(ws, 'D39', detalles)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        if solicitud.nombre_archivo:
            filename = solicitud.nombre_archivo
        else:
            if solicitud.ticket:
                filename = f"{solicitud.ticket.codigo} Solicitud de Pruebas.xlsx"
            else:
                filename = f"Solicitud_{solicitud.id}_{solicitud.fecha_solicitud}.xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        print(f"❌ Error al generar solicitud: {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"Error al generar el archivo: {str(e)}")
        return redirect('extractor:solicitud_detail', id=solicitud.id)