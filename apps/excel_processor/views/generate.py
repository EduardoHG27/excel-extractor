"""
Vistas para generación de documentos Excel (Dictamen, Resultados)
"""
import os
import io
from datetime import datetime
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from extractor.models import Ticket, SolicitudPruebas
from ..utils.helpers import calcular_dias_habiles


full_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)


def restaurar_borde_completo(sheet, celda):
    """Restaura todos los bordes de una celda o rango fusionado"""
    try:
        for merged_range in sheet.merged_cells.ranges:
            if celda in merged_range:
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        sheet.cell(row=row, column=col).border = full_border
                return
        sheet[celda].border = full_border
    except Exception as e:
        print(f"⚠️ No se pudo restaurar borde en {celda}: {e}")


@login_required
def generar_excel_dictamen(request, ticket_id):
    """Genera el Dictamen de Pruebas usando la plantilla"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    solicitud = None
    fecha_solicitud = None
    try:
        solicitud = SolicitudPruebas.objects.filter(ticket=ticket).first()
        if solicitud and solicitud.fecha_solicitud:
            fecha_solicitud = solicitud.fecha_solicitud
    except Exception as e:
        print(f"⚠️ Error al buscar solicitud: {e}")
    
    plantilla_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'plantillas',
        'XXX-XXX-XXX-X-XXX-XXX-XXX DictamenPruebas PRUEBAS.xlsx'
    )
    
    if not os.path.exists(plantilla_path):
        messages.error(request, f"No se encontró la plantilla en: {plantilla_path}")
        return redirect('extractor:ticket_detail', id=ticket.id)
    
    try:
        wb = load_workbook(plantilla_path)
        ws = wb['Dictamen'] if 'Dictamen' in wb.sheetnames else wb.active
        
        def set_cell_value(sheet, coordinate, value):
            try:
                for merged_range in sheet.merged_cells.ranges:
                    if coordinate in merged_range:
                        top_left = merged_range.start_cell.coordinate
                        sheet[top_left] = value
                        return
                sheet[coordinate] = value
            except Exception as e:
                print(f"⚠️ Error en {coordinate}: {e}")
        
        # Desglosar código del ticket
        partes = ticket.codigo.split('-')
        if len(partes) >= 7:
            set_cell_value(ws, 'G2', partes[1])
            set_cell_value(ws, 'I2', partes[2])
            set_cell_value(ws, 'K2', partes[3])
            set_cell_value(ws, 'M2', partes[4])
            set_cell_value(ws, 'O2', partes[5])
            set_cell_value(ws, 'Q2', partes[6])
        
        fecha_actual = datetime.now()
        fecha_actual_str = fecha_actual.strftime('%d/%m/%Y')
        
        # Período de pruebas
        if fecha_solicitud:
            if hasattr(fecha_solicitud, 'strftime'):
                fecha_solicitud_str = fecha_solicitud.strftime('%d/%m/%Y')
            else:
                try:
                    fecha_obj = datetime.strptime(str(fecha_solicitud), '%Y-%m-%d')
                    fecha_solicitud_str = fecha_obj.strftime('%d/%m/%Y')
                except:
                    fecha_solicitud_str = str(fecha_solicitud)
        else:
            fecha_solicitud_str = ticket.fecha_creacion.strftime('%d/%m/%Y') if ticket.fecha_creacion else fecha_actual_str
        
        periodo_pruebas = f"{fecha_solicitud_str} - {fecha_actual_str}"
        set_cell_value(ws, 'K5', periodo_pruebas)
        set_cell_value(ws, 'K6', fecha_actual_str)
        
        # Cálculo de horas
        horas_totales = 0
        if fecha_solicitud:
            if hasattr(fecha_solicitud, 'date'):
                fecha_solicitud_date = fecha_solicitud.date()
            else:
                fecha_solicitud_date = fecha_solicitud
            
            fecha_actual_date = fecha_actual.date()
            dias_habiles = calcular_dias_habiles(fecha_solicitud_date, fecha_actual_date)
            horas_totales = dias_habiles * 8
        
        set_cell_value(ws, 'M18', horas_totales)
        
        # Datos del ticket
        nombre_proyecto = ticket.proyecto.nombre if ticket.proyecto else ''
        if ticket.nombre:
            nombre_limpio = ticket.nombre.replace(" - ", "-")
            partes_nombre = [p.strip() for p in nombre_limpio.split('-') if p.strip()]
            nombre_formateado = " - ".join(partes_nombre)
            proyecto_final = f"{nombre_proyecto} - {nombre_formateado}"
        else:
            proyecto_final = nombre_proyecto
        
        set_cell_value(ws, 'B5', ticket.cliente.nombre if ticket.cliente else '')
        set_cell_value(ws, 'B6', proyecto_final)
        set_cell_value(ws, 'C7', ticket.tipo_servicio.nombre if ticket.tipo_servicio else '')
        
        # Responsable
        responsable_nombre = ""
        if ticket.asignado_a:
            responsable_nombre = ticket.asignado_a.get_full_name() or ticket.asignado_a.username
        elif ticket.creado_por:
            responsable_nombre = ticket.creado_por.get_full_name() or ticket.creado_por.username
        else:
            responsable_nombre = "No asignado"
        
        set_cell_value(ws, 'F28', responsable_nombre)
        
        # Restaurar bordes
        for celda in ['K5', 'K6', 'C9']:
            restaurar_borde_completo(ws, celda)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{ticket.codigo} Dictamen Pruebas.xlsx"'
        
        return response
        
    except Exception as e:
        print(f"❌ Error al generar dictamen: {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"Error al generar dictamen: {str(e)}")
        return redirect('extractor:ticket_detail', id=ticket.id)


@login_required
def generar_excel_resultados(request, ticket_id):
    """Genera el archivo de Documentación de Resultados"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    plantilla_resultados_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'plantillas',
        'XXX-XXX-XXX-X-XXX-XXX-XXX Documentación de Resultados.xlsx'
    )
    
    if os.path.exists(plantilla_resultados_path):
        wb = load_workbook(plantilla_resultados_path)
        ws = wb.active
        ws.title = "Resultados Pruebas"
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados Pruebas"
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 30
    
    ws['C2'] = ticket.codigo
    
    if ws['M3'].value is None or "Versión" not in str(ws['M3'].value):
        ws['M3'] = f"VERSIÓN: Versión {ticket.numero_version or '1.0.0'}"
    
    if ticket.excel_data and ticket.excel_data.detalle_cambios:
        detalle_cambios = ticket.excel_data.detalle_cambios.strip()
        ws['A8'] = detalle_cambios
        ws['A8'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        lineas = detalle_cambios.count('\n') + 1
        altura_estimada = min(lineas * 15, 300)
        ws.row_dimensions[8].height = altura_estimada
    else:
        ws['A8'] = "No se especificaron detalles de cambios."
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{ticket.codigo} Documentación de Resultados.xlsx"'
    
    return response


@login_required
def verificar_plantilla(request):
    """Vista de debug para verificar existencia de plantillas"""
    import os
    from django.conf import settings
    from django.http import HttpResponse
    
    rutas = [
        os.path.join(settings.BASE_DIR, 'static', 'plantillas', 'XXX-XXX-XXX-X-XXX-XXX-XXX DictamenPruebas PRUEBAS.xlsx'),
        os.path.join(settings.BASE_DIR, 'extractor', 'static', 'plantillas', 'XXX-XXX-XXX-X-XXX-XXX-XXX DictamenPruebas PRUEBAS.xlsx'),
        os.path.join(settings.MEDIA_ROOT, 'plantillas', 'XXX-XXX-XXX-X-XXX-XXX-XXX DictamenPruebas PRUEBAS.xlsx'),
    ]
    
    resultado = "<h1>Verificación de Plantilla</h1>"
    resultado += f"<p>BASE_DIR: {settings.BASE_DIR}</p>"
    
    for ruta in rutas:
        existe = os.path.exists(ruta)
        resultado += f"<p>Ruta: {ruta}<br>Existe: {existe}</p>"
        if existe:
            resultado += f"<p>✅ ¡ENCONTRADA AQUÍ!</p>"
    
    return HttpResponse(resultado)