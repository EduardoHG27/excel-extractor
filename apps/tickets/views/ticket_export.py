"""
Vistas para exportación de tickets
"""
import csv
from io import BytesIO
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

from extractor.models import Ticket

logger = logging.getLogger(__name__)


@login_required
def export_tickets_csv_view(request):
    """Exporta tickets a CSV (versión mejorada)"""
    try:
        tickets = Ticket.objects.all().select_related('cliente', 'proyecto', 'tipo_servicio', 'excel_data')
        
        # Aplicar filtros si vienen en GET
        estado = request.GET.get('estado')
        cliente_id = request.GET.get('cliente')
        proyecto_id = request.GET.get('proyecto')
        
        if estado:
            tickets = tickets.filter(estado=estado)
        if cliente_id:
            tickets = tickets.filter(cliente_id=cliente_id)
        if proyecto_id:
            tickets = tickets.filter(proyecto_id=proyecto_id)
        
        response = HttpResponse(content_type='text/csv')
        response.write('\ufeff'.encode('utf-8'))
        
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"tickets_{timestamp}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Código Ticket', 'Estado', 'Cliente', 'Proyecto', 
            'Tipo Servicio', 'Responsable', 'Líder Proyecto', 'Versión',
            'Funcionalidad', 'Detalle Cambios', 'Justificación', 'Fecha Creación'
        ])
        
        for ticket in tickets:
            excel_data = ticket.excel_data
            writer.writerow([
                ticket.id,
                ticket.codigo,
                ticket.get_estado_display(),
                ticket.cliente.nombre if ticket.cliente else '',
                ticket.proyecto.nombre if ticket.proyecto else '',
                ticket.tipo_servicio.nombre if ticket.tipo_servicio else '',
                ticket.responsable_solicitud or '',
                ticket.lider_proyecto or '',
                ticket.numero_version or '',
                excel_data.funcionalidad_liberacion[:100] if excel_data and excel_data.funcionalidad_liberacion else '',
                excel_data.detalle_cambios[:100] if excel_data and excel_data.detalle_cambios else '',
                excel_data.justificacion_cambio[:100] if excel_data and excel_data.justificacion_cambio else '',
                ticket.fecha_creacion.strftime('%d/%m/%Y %H:%M') if ticket.fecha_creacion else ''
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exportando tickets: {str(e)}", exc_info=True)
        messages.error(request, "Error al exportar tickets")
        return redirect('extractor:ticket_list')


@login_required
def export_tickets_excel(request):
    """Exporta los tickets filtrados a un archivo Excel"""
    tickets = Ticket.objects.all().select_related('cliente', 'proyecto', 'tipo_servicio', 'excel_data')
    
    # Aplicar filtros
    estado = request.GET.get('estado')
    cliente_id = request.GET.get('cliente')
    proyecto_id = request.GET.get('proyecto')
    busqueda = request.GET.get('q')
    
    if estado:
        tickets = tickets.filter(estado=estado)
    if cliente_id:
        tickets = tickets.filter(cliente_id=cliente_id)
    if proyecto_id:
        tickets = tickets.filter(proyecto_id=proyecto_id)
    if busqueda:
        tickets = tickets.filter(
            Q(codigo__icontains=busqueda) |
            Q(responsable_solicitud__icontains=busqueda) |
            Q(lider_proyecto__icontains=busqueda)
        )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    headers = [
        'ID', 'Código Ticket', 'Estado', 'Cliente', 'Proyecto', 
        'Tipo Servicio', 'Responsable Solicitud', 'Líder Proyecto',
        'Versión', 'Funcionalidad', 'Detalle Cambios', 'Justificación',
        'Fecha Creación', 'Fecha Actualización'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row, ticket in enumerate(tickets, 2):
        excel_data = ticket.excel_data
        ws.cell(row=row, column=1, value=ticket.id)
        ws.cell(row=row, column=2, value=ticket.codigo)
        ws.cell(row=row, column=3, value=ticket.get_estado_display())
        ws.cell(row=row, column=4, value=ticket.cliente.nombre if ticket.cliente else '')
        ws.cell(row=row, column=5, value=ticket.proyecto.nombre if ticket.proyecto else '')
        ws.cell(row=row, column=6, value=ticket.tipo_servicio.nombre if ticket.tipo_servicio else '')
        ws.cell(row=row, column=7, value=ticket.responsable_solicitud or '')
        ws.cell(row=row, column=8, value=ticket.lider_proyecto or '')
        ws.cell(row=row, column=9, value=ticket.numero_version or '')
        ws.cell(row=row, column=10, value=excel_data.funcionalidad_liberacion if excel_data else '')
        ws.cell(row=row, column=11, value=excel_data.detalle_cambios if excel_data else '')
        ws.cell(row=row, column=12, value=excel_data.justificacion_cambio if excel_data else '')
        ws.cell(row=row, column=13, value=ticket.fecha_creacion.strftime('%d/%m/%Y %H:%M') if ticket.fecha_creacion else '')
        ws.cell(row=row, column=14, value=ticket.fecha_actualizacion.strftime('%d/%m/%Y %H:%M') if ticket.fecha_actualizacion else '')
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"tickets_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response